import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
import shutil
import os
import re


# ================== HÀM LẤY MÃ TỪ TÊN FILE PDF ==================
def extract_codes(filename: str):
    """
    Trả về list các mã dùng để so sánh với Excel.
    - 1–2 ký tự đầu: cho phép A–Z, Đ và cả 0 (vì nhiều người gõ nhầm 0 thay vì O)
    - 5–8 chữ số phía sau
    Nếu prefix có số 0, sẽ tạo thêm phiên bản đổi 0 -> O, để bắt cả 2 trường hợp.
    Ví dụ:
        A012345   -> ['A012345', 'AO12345']
        0Y238103  -> ['0Y238103', 'OY238103']
        CY238103  -> ['CY238103']
    """
    text = (
        filename.upper()
        .replace(" ", "")
        .replace("-", "")
        .replace("_", "")
        .replace(".", "")
    )

    pattern = r"([A-ZĐ0]{1,2})(\d{5,8})"
    codes = set()

    for m in re.finditer(pattern, text):
        prefix_raw = m.group(1)   # có thể chứa 0
        tail = m.group(2)         # toàn số

        # 1) Bản gốc theo đúng tên file
        codes.add(prefix_raw + tail)

        # 2) Nếu prefix có số 0 -> thêm bản đổi 0 -> O
        if "0" in prefix_raw:
            prefix_alt = prefix_raw.replace("0", "O")
            codes.add(prefix_alt + tail)

    return list(codes)


# ================== CHỌN FILE / FOLDER ==================
def chon_file_excel():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not filepath:
        return

    entry_excel.delete(0, tk.END)
    entry_excel.insert(0, filepath)

    try:
        wb = openpyxl.load_workbook(filepath)
        sheet_names = wb.sheetnames
        combo_sheet["values"] = sheet_names
        if sheet_names:
            combo_sheet.current(0)

        # Danh sách cột A..Z cho combobox chọn cột mã
        col_letters = [chr(c) for c in range(ord("A"), ord("Z") + 1)]
        combo_col["values"] = col_letters
        combo_col.set("D")  # mặc định là cột D

    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể đọc file Excel:\n{e}")


def chon_folder_pdf():
    folder = filedialog.askdirectory()
    if folder:
        entry_pdf_folder.delete(0, tk.END)
        entry_pdf_folder.insert(0, folder)


def chon_folder_dich():
    folder = filedialog.askdirectory()
    if folder:
        entry_output_folder.delete(0, tk.END)
        entry_output_folder.insert(0, folder)


def chon_file_luu_excel():
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if filepath:
        entry_excel_output.delete(0, tk.END)
        entry_excel_output.insert(0, filepath)


# ================== HÀM XỬ LÝ CHÍNH ==================
def xu_ly():
    try:
        excel_path = entry_excel.get().strip()
        pdf_folder = entry_pdf_folder.get().strip()
        thu_muc_dich = entry_output_folder.get().strip()
        excel_out = entry_excel_output.get().strip()
        ten_sheet = combo_sheet.get().strip()
        col_letter = combo_col.get().strip().upper()

        # Kiểm tra input
        if not os.path.isfile(excel_path):
            messagebox.showerror("Lỗi", "File Excel không hợp lệ.")
            return
        if not os.path.isdir(pdf_folder):
            messagebox.showerror("Lỗi", "Thư mục PDF không hợp lệ.")
            return
        if not os.path.isdir(thu_muc_dich):
            messagebox.showerror("Lỗi", "Thư mục đích không hợp lệ.")
            return
        if not ten_sheet:
            messagebox.showerror("Lỗi", "Vui lòng chọn sheet.")
            return
        if not col_letter:
            messagebox.showerror("Lỗi", "Vui lòng chọn cột chứa mã.")
            return

        wb = openpyxl.load_workbook(excel_path)
        if ten_sheet not in wb.sheetnames:
            messagebox.showerror("Lỗi", "Sheet được chọn không tồn tại trong Excel.")
            return

        sheet = wb[ten_sheet]
        col_index = column_index_from_string(col_letter)

        # ========== 1. Đọc mã từ Excel ==========
        codes_set = set()      # tập mã để so sánh
        code_to_rows = {}      # mã -> các dòng chứa mã đó (để tô màu)

        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col_index).value
            if not cell_value:
                continue

            text_raw = str(cell_value).upper()
            # Bỏ khoảng trắng, -, _, . để "CY 238103", "CY-238103" -> "CY238103"
            text = (
                text_raw
                .replace(" ", "")
                .replace("-", "")
                .replace("_", "")
                .replace(".", "")
            )

            # Prefix trong Excel chỉ cho phép chữ (A-Z, Đ), không chấp nhận 0 ở đây
            pattern = r"([A-ZĐ]{1,2})(\d{5,8})"
            found_codes = []

            for m in re.finditer(pattern, text):
                prefix = m.group(1)
                tail = m.group(2)
                code = prefix + tail
                found_codes.append(code)

            if found_codes:
                for code in found_codes:
                    codes_set.add(code)
                    code_to_rows.setdefault(code, set()).add(row)

        if not codes_set:
            messagebox.showwarning("Cảnh báo", "Không tìm thấy mã hợp lệ trong Excel.")
            return

        print(f"Excel có {len(codes_set)} mã hợp lệ.")

        # ========== 2. Duyệt file PDF ==========
        file_copied_count = 0
        codes_found = set()  # các mã thực sự đã xuất hiện trong file PDF

        for root_dir, dirs, files in os.walk(pdf_folder):
            for file in files:
                if not file.lower().endswith(".pdf"):
                    continue

                file_codes = extract_codes(file)  # list mã (cả bản 0 và O nếu có)
                if not file_codes:
                    continue

                # Nếu BẤT KỲ mã nào trong file trùng với Excel -> copy
                if any(code in codes_set for code in file_codes):
                    src = os.path.join(root_dir, file)
                    dst = os.path.join(thu_muc_dich, file)  # giữ nguyên tên

                    shutil.copy2(src, dst)
                    file_copied_count += 1

                    # cập nhật codes_found để tô màu Excel
                    for code in file_codes:
                        if code in codes_set:
                            codes_found.add(code)

                    print(f"✔ Copy: {file} | Mã trong file: {file_codes}")

        # ========== 3. Tô màu Excel (nếu có chọn nơi lưu) ==========
        if excel_out:
            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            rows_to_color = set()
            for code in codes_found:
                rows_to_color.update(code_to_rows.get(code, []))

            for row in rows_to_color:
                for cell in sheet[row]:
                    cell.fill = fill

            wb.save(excel_out)
            print(f"Đã lưu Excel: {excel_out}")

        messagebox.showinfo(
            "Xong",
            f"Đã copy {file_copied_count} file PDF.\n"
            f"Số mã khớp: {len(codes_found)}"
        )

    except Exception as e:
        messagebox.showerror("Lỗi", str(e))


# ================== GIAO DIỆN TKINTER ==================
root = tk.Tk()
root.title("Lọc file PDF theo mã trong Excel")

# Hàng 0: File Excel
tk.Label(root, text="File Excel:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
entry_excel = tk.Entry(root, width=60)
entry_excel.grid(row=0, column=1, padx=5, pady=5)
tk.Button(root, text="Browse", command=chon_file_excel)\
    .grid(row=0, column=2, padx=5, pady=5)

# Hàng 1: Sheet
tk.Label(root, text="Sheet:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
combo_sheet = ttk.Combobox(root, width=57, state="readonly")
combo_sheet.grid(row=1, column=1, columnspan=2, sticky="w", padx=5, pady=5)

# Hàng 2: Cột mã
tk.Label(root, text="Cột chứa mã:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
combo_col = ttk.Combobox(root, width=57, state="readonly")
combo_col.grid(row=2, column=1, columnspan=2, sticky="w", padx=5, pady=5)

# Hàng 3: Thư mục PDF
tk.Label(root, text="Thư mục PDF:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
entry_pdf_folder = tk.Entry(root, width=60)
entry_pdf_folder.grid(row=3, column=1, padx=5, pady=5)
tk.Button(root, text="Browse", command=chon_folder_pdf)\
    .grid(row=3, column=2, padx=5, pady=5)

# Hàng 4: Thư mục đích
tk.Label(root, text="Thư mục đích:").grid(row=4, column=0, sticky="e", padx=5, pady=5)
entry_output_folder = tk.Entry(root, width=60)
entry_output_folder.grid(row=4, column=1, padx=5, pady=5)
tk.Button(root, text="Browse", command=chon_folder_dich)\
    .grid(row=4, column=2, padx=5, pady=5)

# Hàng 5: Lưu Excel sau tô màu (tuỳ chọn)
tk.Label(root, text="Lưu Excel sau tô màu (có thể bỏ trống):").grid(
    row=5, column=0, sticky="e", padx=5, pady=5
)
entry_excel_output = tk.Entry(root, width=60)
entry_excel_output.grid(row=5, column=1, padx=5, pady=5)
tk.Button(root, text="Browse", command=chon_file_luu_excel)\
    .grid(row=5, column=2, padx=5, pady=5)

# Hàng 6: Nút chạy
tk.Button(
    root, text="Chạy xử lý", bg="green", fg="white", height=2, command=xu_ly
).grid(row=6, column=1, pady=20)

root.mainloop()
