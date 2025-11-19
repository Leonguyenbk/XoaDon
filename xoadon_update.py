import time, traceback, threading, sys, json, re, os
from venv import logger
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, ElementClickInterceptedException, JavascriptException,
    StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException,
    NoSuchWindowException
)

# ---- Tkinter GUI ----
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# ---- Excel Import ----
try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    openpyxl = None

# ============== LOG UI HELPERS ==============
class UILogger:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def log(self, msg):
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode(sys.stdout.encoding, errors='replace').decode(sys.stdout.encoding))
        if self.text_widget:
            self.text_widget.after(0, lambda: self._append(msg))

    def _append(self, msg):
        self.text_widget.configure(state="normal")
        self.text_widget.insert("end", msg + "\n")
        self.text_widget.see("end")
        self.text_widget.configure(state="disabled")

# ============== WAITERS / HELPERS ==============
def wait_xuly_modal(driver, timeout=20):
    wait = WebDriverWait(driver, timeout)
    driver.switch_to.default_content()
    modal = wait.until(EC.visibility_of_element_located((
        By.CSS_SELECTOR, "div.modal.modal-fullscreen.in[id^='mdlXuLyDonDangKy-'][style*='display: block']"
    )))
    try:
        WebDriverWait(driver, 5).until(lambda d: d.execute_script("return (window.jQuery? jQuery.active:0)") == 0)
    except Exception:
        pass
    return modal

def wait_jstree_ready_in(container_el, timeout=20):
    end = time.time() + timeout
    while time.time() < end:
        trees = container_el.find_elements(By.CSS_SELECTOR, "#treeDonDangKy")
        if trees:
            anchors = trees[0].find_elements(By.CSS_SELECTOR, "a.jstree-anchor")
            if anchors:
                if not (len(anchors) == 1 and "Không có dữ liệu" in (anchors[0].text or "")):
                    return trees[0]
        time.sleep(0.2)
    raise TimeoutException("jsTree chưa có dữ liệu trong thời gian cho phép.")

def find_tt_dangky_anchor(tree_el):
    xpaths = [
        ".//a[.//b[normalize-space()='Thông tin đăng ký']]",
        ".//a[normalize-space()='Thông tin đăng ký']",
        ".//a[contains(normalize-space(.), 'Thông tin đăng ký')]",
    ]
    for xp in xpaths:
        els = tree_el.find_elements(By.XPATH, xp)
        if els:
            return els[0]
    raise NoSuchElementException("Không tìm thấy anchor 'Thông tin đăng ký' trong jsTree.")

def wait_page_idle(driver, wait, extra_ms=300):
    wait.until(lambda x: x.execute_script("return document.readyState") == "complete")
    time.sleep(extra_ms/1000.0)

def row_is_highlighted(row):
    """
    True nếu dòng đã tô fill (fill_type khác None/'none').
    """
    for cell in row:
        fill = cell.fill
        if fill is not None and fill.fill_type not in (None, "none"):
            return True
    return False

def switch_to_iframe_containing_table(driver, table_id="tblTTThuaDat", timeout=10):
    driver.switch_to.default_content()
    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    deadline = time.time() + timeout
    for idx in range(len(iframes)):
        if time.time() > deadline:
            break
        driver.switch_to.default_content()
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        try:
            driver.switch_to.frame(iframes[idx])
            if driver.find_elements(By.CSS_SELECTOR, f"#{table_id}"):
                return True
            inner_iframes = driver.find_elements(By.TAG_NAME, "iframe")
            for j in range(len(inner_iframes)):
                driver.switch_to.frame(inner_iframes[j])
                if driver.find_elements(By.CSS_SELECTOR, f"#{table_id}"):
                    return True
                driver.switch_to.parent_frame()
        except Exception:
            continue
    driver.switch_to.default_content()
    return False

def wait_for_table_loaded(driver, table_id="tblTTThuaDat", timeout=15):
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, f"#{table_id}_processing"))
        )
    except TimeoutException:
        pass

def safe_click_row_css(driver, wait, row_css="#tblTraCuuDotBanGiao tbody tr", logger=None):
    wait_page_idle(driver, wait, 300)
    row = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, row_css)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row)
    cell = row.find_element(By.CSS_SELECTOR, "td:nth-child(2)")
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//table[@id='tblTraCuuDotBanGiao']//tbody//tr[1]//td[2]")))
        cell.click()
        return
    except ElementClickInterceptedException:
        wait_page_idle(driver, wait, 300)
        try:
            cell.click()
            return
        except ElementClickInterceptedException:
            pass
    try:
        driver.execute_script("""
            document.querySelectorAll('.jquery-loading-modal__bg')
                  .forEach(el => { el.style.pointerEvents='none'; el.style.display='none'; });
        """)
    except JavascriptException:
        pass
    try:
        driver.execute_script("arguments[0].click();", cell)
        return
    except Exception:
        pass
    first_cell = row.find_element(By.CSS_SELECTOR, "td:nth-child(1)")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", first_cell)
    driver.execute_script("arguments[0].click();", first_cell)

def goto_page(driver, page_number, table_id="tblTTThuaDat", verify_timeout=5):
    driver.execute_script(f"""
        if (window.jQuery && jQuery.fn.DataTable) {{
            var table = jQuery('#{table_id}').DataTable();
            var info  = table.page.info();
            var maxp  = info.pages || 1;
            var target = Math.max(0, Math.min({page_number}-1, maxp-1));
            table.page(target).draw('page');
        }}
    """)
    end = time.time() + verify_timeout
    target0 = max(0, page_number-1)
    while time.time() < end:
        ok = driver.execute_script(f"""
            try {{
                var t = jQuery('#{table_id}').DataTable();
                return t.page.info().page;
            }} catch(e){{ return -1; }}
        """)
        if ok == target0:
            return True
        time.sleep(0.2)
    return False

def go_next_datatables(driver, table_id="tblTTThuaDat", timeout=15):
    wait = WebDriverWait(driver, timeout)
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, f"#{table_id}_processing")))
    except TimeoutException:
        pass
    li_next = wait.until(EC.presence_of_element_located((By.ID, f"{table_id}_next")))
    if "disabled" in (li_next.get_attribute("class") or ""):
        return False
    a_next = li_next.find_element(By.TAG_NAME, "a")
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"#{table_id}_next a")))
    w, h, vis = driver.execute_script("""
        const a = arguments[0];
        const r = a.getBoundingClientRect();
        const style = window.getComputedStyle(a);
        return [r.width, r.height, style.visibility !== 'hidden' && style.display !== 'none'];
    """, a_next)
    if not (w > 0 and h > 0 and vis):
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", a_next)
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"#{table_id}_next a")))
    first_row = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"#{table_id} tbody tr")))
    try:
        a_next.click()
    except Exception:
        driver.execute_script("arguments[0].click();", a_next)
    try:
        wait.until(EC.staleness_of(first_row))
    except (TimeoutException, StaleElementReferenceException):
        try:
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, f"#{table_id}_processing")))
        except TimeoutException:
            pass
    return True

def handle_whole_page_action(driver, logger: UILogger, table_id="tblTTThuaDat", timeout=15):
    wait = WebDriverWait(driver, timeout)
    wait.until(EC.presence_of_element_located((By.ID, table_id)))
    rows = driver.find_elements(By.CSS_SELECTOR, f"#{table_id} tbody > tr:not(.child)")

    visible_rows = []
    for r in rows:
        try:
            tds = r.find_elements(By.CSS_SELECTOR, "td")
            if tds and r.is_displayed():
                visible_rows.append((r, tds))
        except StaleElementReferenceException:
            continue

    if len(visible_rows) < 1:
        logger.log("   (Không có hàng nào hiển thị để chọn)")
        return 0

    first_row, first_tds = visible_rows[0]
    last_row, last_tds = visible_rows[-1]

    def pick_click_target(row, tds):
        for css in ["input[type='checkbox']:not([disabled])", "button", "a"]:
            try:
                el = row.find_element(By.CSS_SELECTOR, css)
                if el.is_displayed():
                    return el
            except NoSuchElementException:
                pass
        return tds[0]

    first_target = pick_click_target(first_row, first_tds)
    last_target = pick_click_target(last_row, last_tds)

    def ensure_visible_and_sized(el):
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        WebDriverWait(driver, timeout).until(lambda d: d.execute_script("""
            const r = arguments[0].getBoundingClientRect();
            const s = getComputedStyle(arguments[0]);
            return r.width > 0 && r.height > 0 && s.display!=='none' && s.visibility!=='hidden';
        """, el))

    try:
        ensure_visible_and_sized(first_target)
        first_target.click()
        if len(visible_rows) > 1:
            ensure_visible_and_sized(last_target)
            ActionChains(driver).key_down(Keys.SHIFT).click(last_target).key_up(Keys.SHIFT).perform()
    except Exception as e:
        logger.log(f"   (Lỗi Shift-Click, thử fallback... Lỗi: {e})")
        for row, tds in visible_rows:
            try:
                target = pick_click_target(row, tds)
                ensure_visible_and_sized(target)
                target.click()
            except Exception:
                continue

    logger.log("   → Đã chọn tất cả, bắt đầu lọc bỏ những bản ghi đã duyệt...")
    time.sleep(0.2)

    actions = ActionChains(driver).key_down(Keys.CONTROL)
    deselected_count = 0
    selected_rows = driver.find_elements(By.CSS_SELECTOR, f"#{table_id} tbody tr.selected")
    for row in selected_rows:
        try:
            txt = (row.get_attribute("innerText") or row.text).strip().lower()
            if "đã duyệt ghi đè" in txt:
                actions.click(row.find_element(By.CSS_SELECTOR, "td:first-child"))
                deselected_count += 1
        except (StaleElementReferenceException, NoSuchElementException):
            continue
    actions.key_up(Keys.CONTROL).perform()

    selected_count = driver.execute_script(f"""
        try {{
            if (window.jQuery && jQuery.fn.DataTable) {{
                const dt = jQuery("#{table_id}").DataTable();
                return dt.rows({{selected:true, page:'current'}}).count();
            }}
        }} catch(e) {{}}
        const table = document.querySelector("#{table_id}");
        return table ? table.querySelectorAll("tbody tr.selected").length : 0;
    """)

    if deselected_count > 0:
        logger.log(f"   → Đã bỏ chọn {deselected_count} bản ghi đã duyệt. Còn lại {selected_count} bản ghi.")

    return selected_count

def quick_confirm_if_present(driver, root_el=None, soft_timeout=1.2):
    try:
        scope = root_el if root_el is not None else driver

        btns = scope.find_elements(By.CSS_SELECTOR, ".swal2-container .swal2-confirm")
        if not btns:
            btns = scope.find_elements(By.CSS_SELECTOR, ".modal.in .btn-primary, .modal.show .btn-primary")

        if not btns:
            xp = ".//button[normalize-space()='Đồng ý' or normalize-space()='Xác nhận' or normalize-space()='OK' or normalize-space()='Có' or normalize-space()='Yes']"
            try:
                btns = scope.find_elements(By.XPATH, xp)
            except Exception:
                btns = []

        if not btns:
            return False

        cand = None
        for b in btns:
            try:
                vis = driver.execute_script("""
                    const el = arguments[0];
                    const r = el.getBoundingClientRect();
                    const s = getComputedStyle(el);
                    return r.width>0 && r.height>0 && s.visibility!=='hidden' && s.display!=='none';
                """, b)
                if vis:
                    cand = b
                    break
            except Exception:
                continue
        if cand is None:
            return False

        try:
            driver.execute_script("""
                document.querySelectorAll('.modal-backdrop, .swal2-container, .jquery-loading-modal__bg')
                    .forEach(el=>{ el.style.pointerEvents='auto'; });
            """)
        except Exception:
            pass

        try:
            cand.click()
            return True
        except Exception:
            pass

        try:
            driver.execute_script("arguments[0].click();", cand)
            return True
        except Exception:
            pass

        try:
            driver.switch_to.active_element.send_keys(Keys.ENTER)
            return True
        except Exception:
            pass

        return False
    except Exception:
        return False

def wait_processing_quick(driver, table_id="tblTTThuaDat", max_wait=6):
    def cond(d):
        try:
            ajax_zero = d.execute_script("return (window.jQuery ? jQuery.active : 0)") == 0
            proc = d.execute_script(f"""
                var e = document.querySelector('#{table_id}_processing');
                if(!e) return true;
                var s = getComputedStyle(e);
                return (s.display==='none' || s.visibility==='hidden' || e.offsetParent===null);
            """)
            return ajax_zero and proc
        except Exception:
            return True
    try:
        WebDriverWait(driver, max_wait, poll_frequency=0.1).until(cond)
        return True
    except Exception:
        return False

def hard_jump_pagination(driver, page_number, table_id="tblTTThuaDat", timeout=10):
    wait = WebDriverWait(driver, timeout)
    cur = driver.execute_script(f"""
        try {{
            return jQuery('#{table_id}').DataTable().page.info().page + 1;
        }} catch(e) {{ return 1; }}
    """) or 1

    if page_number == cur:
        return True

    try:
        btn = wait.until(EC.presence_of_element_located((
            By.XPATH, f"//div[@id='{table_id}_paginate']//a[normalize-space(text())='{page_number}']"
        )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
    except TimeoutException:
        step = 1 if page_number > cur else -1
        next_sel = f"#{table_id}_next a"
        prev_sel = f"#{table_id}_previous a"
        while cur != page_number:
            sel = next_sel if step == 1 else prev_sel
            try:
                a = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
                a.click()
            except Exception:
                driver.execute_script("document.querySelector(arguments[0])?.click()", sel)
            wait_for_table_loaded(driver, table_id, timeout=10)
            cur = driver.execute_script(f"return jQuery('#{table_id}').DataTable().page.info().page + 1;") or cur
            if cur == page_number:
                break

    wait_for_table_loaded(driver, table_id, timeout=10)
    cur2 = driver.execute_script(f"return jQuery('#{table_id}').DataTable().page.info().page + 1;")
    return cur2 == page_number

def all_jconfirm_closed(driver):
    modals = driver.find_elements(By.CSS_SELECTOR, ".jconfirm-scrollpane")
    if not modals:
        return True
    for m in modals:
        try:
            if m.is_displayed():
                return False
        except Exception:
            continue
    return True

def wait_all_jconfirm_closed(driver, timeout=15):
    try:
        WebDriverWait(driver, timeout).until(lambda d: all_jconfirm_closed(d))
    except TimeoutException:
        pass

def switch_to_frame_having(driver, by, value, timeout=8):
    driver.switch_to.default_content()
    try:
        if driver.find_elements(by, value):
            return True
    except Exception:
        pass
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    deadline = time.time() + timeout
    for i in range(len(frames)):
        if time.time() > deadline:
            break
        driver.switch_to.default_content()
        frames = driver.find_elements(By.TAG_NAME, "iframe")
        try:
            driver.switch_to.frame(frames[i])
            if driver.find_elements(by, value):
                return True
            inner = driver.find_elements(By.TAG_NAME, "iframe")
            for j in range(len(inner)):
                driver.switch_to.frame(inner[j])
                if driver.find_elements(by, value):
                    return True
                driver.switch_to.parent_frame()
        except Exception:
            continue
    driver.switch_to.default_content()
    return False

def wait_tracuu_module_ready(driver, timeout=60):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#donDangKyTraCuuModule"))
    )
    WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#donDangKyTraCuuModule"))
    )
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".jquery-loading-modal__bg"))
        )
    except:
        pass
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let el = document.querySelector("#donDangKyTraCuuModule");
            if (!el) return false;
            return el.offsetHeight > 0 && el.offsetWidth > 0;
        """)
    )
    print("✅ Module tra cứu (#donDangKyTraCuuModule) đã load xong!")

def wait_tracuu_section_ready(driver, timeout=60):
    selector = "#donDangKyTraCuuModule > div.panel-body > div > div:nth-child(3)"

    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
    )
    WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
    )
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".jquery-loading-modal__bg"))
        )
    except:
        pass
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let el = document.querySelector(arguments[0]);
            if (!el) return false;
            let rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        """, selector)
    )
    print("✅ Vùng tra cứu (div:nth-child(3)) đã load xong!")

def wait_and_count_tblTraCuu(driver, timeout=60):
    table_selector = "#tblTraCuuTinhHinhDangKy"

    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, table_selector))
    )

    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".jquery-loading-modal__bg"))
        )
    except:
        pass

    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let p = document.querySelector("#tblTraCuuTinhHinhDangKy_processing");
            if (p && p.offsetParent !== null) return false;
            return true;
        """)
    )

    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let tb = document.querySelector("#tblTraCuuTinhHinhDangKy tbody");
            return tb && tb.children.length >= 0;
        """)
    )

    count = driver.execute_script("""
        let table = document.querySelector("#tblTraCuuTinhHinhDangKy");
        if (!table) return -1;

        let rows = table.querySelectorAll("tbody tr");
        if (!rows || rows.length === 0) return 0;

        let count = 0;
        rows.forEach(r => {
            let td = r.querySelector("td");
            if (td && td.classList.contains("dataTables_empty")) return; 
            count++;
        });

        return count;
    """)

    print("➡️ Số bản ghi:", count)
    return count

def wait_query_done(driver, timeout=30, ajax_wait=5):
    end_time = time.time() + timeout

    try:
        WebDriverWait(driver, 5).until(
            lambda d: d.execute_script("return window.jQuery !== undefined;")
        )
    except Exception:
        return

    phase1_end = time.time() + ajax_wait
    saw_ajax = False
    while time.time() < phase1_end:
        try:
            active = driver.execute_script("return jQuery.active;")
            if active > 0:
                saw_ajax = True
                break
        except Exception:
            break
        time.sleep(0.1)

    if not saw_ajax:
        return

    while time.time() < end_time:
        try:
            active = driver.execute_script("return jQuery.active;")
            if active == 0:
                return
        except Exception:
            return
        time.sleep(0.1)

def wait_query_xoadon(driver, timeout=30, ajax_wait=5, max_after_first_ajax=10):
    try:
        WebDriverWait(driver, 3).until(
            lambda d: d.execute_script("return window.jQuery !== undefined;")
        )
    except Exception:
        return

    phase1_end = time.time() + ajax_wait
    saw_ajax = False

    while time.time() < phase1_end:
        try:
            active = driver.execute_script("return jQuery.active;")
            if active > 0:
                saw_ajax = True
                break
        except Exception:
            return
        time.sleep(0.1)

    if not saw_ajax:
        return

    phase2_end = time.time() + max_after_first_ajax
    THRESH = 1
    stable_required = 5
    stable_count = 0

    while time.time() < phase2_end:
        try:
            active = driver.execute_script("return jQuery.active;")
        except Exception:
            return

        if active <= THRESH:
            stable_count += 1
            if stable_count >= stable_required:
                return
        else:
            stable_count = 0

        time.sleep(0.1)
    return

def retry_delete_confirm_if_jconfirm(driver, wait, logger: UILogger = None):
    try:
        modals = driver.find_elements(By.CSS_SELECTOR, "div.jconfirm-scrollpane")
        has_open = False
        for m in modals:
            try:
                if m.is_displayed():
                    has_open = True
                    break
            except Exception:
                continue

        if not has_open:
            return False

        if logger:
            logger.log("⚠️ Phát hiện popup jConfirm đang che, thử ĐỒNG Ý/OK lại...")

        try:
            dongy_selector = "div.jconfirm.jconfirm-open .jconfirm-buttons button.btn.btn-orange"
            dongy_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, dongy_selector))
            )

            print("👉 Popup xác nhận đã hiện (retry), nhấn ĐỒNG Ý")
            driver.execute_script("arguments[0].click();", dongy_btn)

        except Exception as e:
            print(f"❌ Không thấy hoặc không click được nút ĐỒNG Ý (retry): {e}")
            if logger:
                logger.log("❌ Không thấy popup ĐỒNG Ý khi retry.")
            try:
                driver.switch_to.active_element.send_keys(Keys.ENTER)
                print("⌨️ Đã thử nhấn ENTER để xác nhận (retry).")
            except Exception as enter_e:
                print(f" Lỗi khi thử nhấn ENTER (retry): {enter_e}")
                return False

        wait_query_xoadon(driver, timeout=30)
        wait_all_jconfirm_closed(driver, timeout=10)

        try:
            ok_selector = "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open .jconfirm-buttons button"
            ok_wait = WebDriverWait(driver, 30)
            ok_btn = ok_wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ok_selector))
            )

            print("👉 Popup thông báo thành công (retry) đã hiện, nhấn OK")
            driver.execute_script("arguments[0].click();", ok_btn)

        except Exception as e:
            print(f"ℹ️ Không thấy popup 'OK' sau khi xóa (retry), hoặc đã tự đóng: {e}")

        wait_all_jconfirm_closed(driver, timeout=10)

        if logger:
            logger.log("✅ Đã xử lý lại popup jConfirm (Đồng ý/OK).")

        return True

    except Exception as e:
        print(f"❌ Lỗi khi retry popup jConfirm: {e}")
        if logger:
            logger.log(f"❌ Lỗi khi retry popup jConfirm: {e}")
        return False

def chon_ban_ghi_dau_tien(driver, timeout=30, logger: UILogger = None):
    wait = WebDriverWait(driver, timeout)

    first_row = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr")
        )
    )

    if "Không tìm thấy" in first_row.text:
        return False

    checkbox = wait.until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr:nth-child(1) td.select-checkbox")
        )
    )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", checkbox)

    try:
        checkbox.click()
    except ElementClickInterceptedException as e:
        if "jconfirm-scrollpane" in str(e):
            if logger:
                logger.log("⚠️ Click checkbox bị chặn bởi popup jConfirm. Đang retry ĐỒNG Ý/OK...")

            handled = retry_delete_confirm_if_jconfirm(driver, wait, logger)

            if not handled:
                raise

            if logger:
                logger.log("🔁 Thử click lại checkbox dòng đầu sau khi xử lý popup...")

            first_row = wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr")
                )
            )
            if "Không tìm thấy" in first_row.text:
                return False

            checkbox = wait.until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr:nth-child(1) td.select-checkbox")
                )
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", checkbox)
            driver.execute_script("arguments[0].click();", checkbox)
        else:
            raise

    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr.selected")
        )
    )

    btn_chon = wait.until(
        EC.element_to_be_clickable((By.ID, "btnLuuChonTinhHinhDangKy"))
    )
    btn_chon.click()

    wait.until(
        EC.invisibility_of_element_located((By.ID, "donDangKyTraCuuModule"))
    )

    return True

def click_step_GiayChungNhan(driver, timeout=30):
    wait = WebDriverWait(driver, timeout)

    selector = "li.tour_kekhaidangky_step16"

    step = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))

    if "active" in step.get_attribute("class"):
        return True

    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector))).click()

    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, selector + ".clickable.active")
        )
    )

    return True

def kiem_tra_tree_gcn(driver):
    anchors = driver.find_elements(By.CSS_SELECTOR, "#treeGiayChungNhan a.jstree-anchor")

    if not anchors:
        print("❌ Không có dữ liệu trong #treeGiayChungNhan")
        return ("no_data", None, "")

    text = anchors[0].text.strip()
    text_lower = text.lower()

    if "không có dữ liệu" in text_lower or "không có giữ liệu" in text_lower:
        print("ℹ️ Cây GCN hiển thị 'Không có dữ liệu'")
        return ("no_data", None, text)

    pattern = r"Số phát hành:\s*((?:[A-Za-zĐđ]{1,2}\s?\d{5,8})|(?:\d{5,8}))"
    match = re.search(pattern, text)

    if match:
        gcn_code = match.group(1).strip()
        print(f"✅ Có dữ liệu GCN, Số phát hành: {gcn_code}")
        return ("has_gcn", gcn_code, text)
    else:
        print("✅ Có dữ liệu trong cây GCN nhưng không tìm thấy 'Số phát hành'")
        return ("has_data", None, text)

def perform_bo_don(driver, wait, logger: UILogger, reason="", so_to=None, so_thua=None, gcn_code=None):
    log_message = f"✅ {reason} Tiến hành bỏ đơn..."
    logger.log(log_message)

    if ("GCN" in reason or "thế chấp" in reason) and so_to and so_thua:
        try:
            with open("thua_dat_co_gcn.txt", "a", encoding="utf-8") as f:
                f.write(f"Số tờ: {so_to}, Số thửa: {so_thua}, Mã GCN: {gcn_code or 'N/A'}\n")
            logger.log(f"💾 Đã lưu thông tin thửa đất có GCN vào file 'thua_dat_co_gcn.txt'.")
        except Exception as e:
            logger.log(f"⚠️ Lỗi khi ghi file txt: {e}")
            print(f"⚠️ Lỗi khi ghi file txt: {e}")
    
    try:
        btn_bo_don = wait.until(EC.element_to_be_clickable((By.ID, "btnBoDonDangKy")))
        btn_bo_don.click()

        wait.until(
            EC.visibility_of_element_located((
                By.CSS_SELECTOR,
                "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open"
            ))
        )
        print("👉 Popup xác nhận 'Bỏ đơn' đã xuất hiện")

        btn_orange = wait.until(
            EC.element_to_be_clickable((
                By.CSS_SELECTOR,
                "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open .jconfirm-buttons button.btn.btn-orange"
            ))
        )
        print("👉 Nút cam 'Đồng ý' đã sẵn sàng")

        btn_orange.click()
        print("👉 Đã nhấn nút cam 'Đồng ý'")

        wait.until(
            EC.invisibility_of_element_located((
                By.CSS_SELECTOR,
                "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open"
            ))
        )
        print("✅ Popup 'Bỏ đơn' đã đóng")

        WebDriverWait(driver, 15).until(lambda d: all_jconfirm_closed(d))
        print("✅ Tất cả popup đã đóng – Bỏ đơn thành công!")
        logger.log("✅ Thao tác 'Bỏ đơn' hoàn tất.")
        return True

    except Exception as e:
        logger.log(f"❌ Lỗi trong quá trình 'Bỏ đơn': {e}")
        print(f"❌ Lỗi trong quá trình 'Bỏ đơn': {e}")
        return True

def search_and_process_plot(driver, wait, logger: UILogger, so_to, so_thua):
    try:
        so_thua_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,
            "#dvTraCuuTinhHinhDangKyChiTiet > div:nth-child(2) > div.col-md-8.col-sm-12 > fieldset > div:nth-child(2) > div:nth-child(1) > div > input"
        )))
        so_thua_input.clear()
        so_thua_input.send_keys(so_thua)

        so_to_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,
            "#dvTraCuuTinhHinhDangKyChiTiet > div:nth-child(2) > div.col-md-8.col-sm-12 > fieldset > div:nth-child(2) > div:nth-child(2) > div > input"
        )))
        so_to_input.clear()
        so_to_input.send_keys(so_to)

        so_thua_input.send_keys(Keys.ENTER)

        wait_tracuu_section_ready(driver, timeout=60)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "tblTraCuuTinhHinhDangKy_info"))
        )
        wait_query_done(driver, timeout=60)
        so_ban_ghi = wait_and_count_tblTraCuu(driver)
        logger.log(f"✅ Đã nhập Số tờ: {so_to}, Số thửa: {so_thua}. Số bản ghi tìm được: {so_ban_ghi}.")

        if so_ban_ghi == 0:
            logger.log("❌ Không tìm thấy bản ghi nào. Tìm thửa tiếp theo...")
            return False, "Không tìm thấy bản ghi"

        chon_ban_ghi_dau_tien(driver, timeout=30, logger=logger)
        wait_query_done(driver, timeout=60)
        click_step_GiayChungNhan(driver, timeout=30)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "treeGiayChungNhan"))
        )
        status, gcn_code, raw_text = kiem_tra_tree_gcn(driver)
        if status == "no_data":
            logger.log("👉 Kết quả kiểm tra GCN: Không có dữ liệu (sẽ XÓA ĐƠN).")
        elif status == "has_gcn":
            logger.log(f"👉 Kết quả kiểm tra GCN: Có mã GCN {gcn_code} (sẽ BỎ ĐƠN).")
        else:
            logger.log("👉 Kết quả kiểm tra GCN: Có dữ liệu nhưng không có 'Số phát hành' (sẽ BỎ ĐƠN).")

        if status == "no_data":
            try:
                btn_xoa = wait.until(EC.element_to_be_clickable((By.ID, "btnXoaDonDangKy")))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_xoa)
                btn_xoa.click()
                print("👉 Đã nhấn nút Xóa đơn đăng ký")
            except Exception as e:
                print(f"❌ Không tìm thấy hoặc không click được nút Xóa đơn đăng ký: {e}")
                logger.log("❌ Không tìm thấy nút Xóa đơn đăng ký.")
                return True, "Lỗi khi xử lý thửa (không click được nút Xóa đơn đăng ký)"

            WebDriverWait(driver, 15).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR,
                    "div.jconfirm.jconfirm-open .jconfirm-scrollpane")))

            try:
                dongy_selector = "div.jconfirm.jconfirm-open .jconfirm-buttons button.btn.btn-orange"
                dongy_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, dongy_selector)))
                
                print("👉 Popup xác nhận đã hiện, nhấn ĐỒNG Ý")

                driver.execute_script("arguments[0].click();", dongy_btn)

            except Exception as e:
                print(f"❌ Không thấy hoặc không click được nút ĐỒNG Ý: {e}")
                logger.log("❌ Không thấy popup xác nhận khi xóa.")
                try:
                    driver.switch_to.active_element.send_keys(Keys.ENTER)
                    print("⌨️ Đã thử nhấn ENTER để xác nhận.")
                except Exception as enter_e:
                    print(f" Lỗi khi thử nhấn ENTER: {enter_e}")
                    return True, "Lỗi khi xử lý thửa (không click được Đồng ý khi xóa)"

            wait_query_xoadon(driver, timeout=60)
            wait_all_jconfirm_closed(driver, timeout=15)

            try:
                ok_selector = "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open .jconfirm-buttons button"
                ok_wait = WebDriverWait(driver, 60)
                ok_btn = ok_wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ok_selector)))
                
                print("👉 Popup thông báo thành công đã hiện, nhấn OK")
                driver.execute_script("arguments[0].click();", ok_btn)

            except Exception as e:
                print(f"ℹ️ Không thấy popup 'OK' sau khi xóa, hoặc đã tự đóng: {e}")
            
            wait_all_jconfirm_closed(driver, timeout=15)

            try:
                WebDriverWait(driver, 15).until(lambda d: all_jconfirm_closed(d))
            except Exception:
                print("⚠ Popup không biến mất đúng hạn, nhưng có thể đã xử lý xong")
                logger.log("⚠ Thao tác xóa hoàn tất nhưng popup không tự đóng.")

            return True, "Đã xóa đơn"

        else:
            if status == "has_gcn":
                success = perform_bo_don(
                    driver, wait, logger,
                    reason="Thửa đất đã có GCN.",
                    so_to=so_to, so_thua=so_thua,
                    gcn_code=gcn_code
                )
                note = f"bỏ đơn do có mã GCN {gcn_code}"
            else:
                success = perform_bo_don(
                    driver, wait, logger,
                    reason="Thửa đất có dữ liệu GCN (không có số phát hành).",
                    so_to=so_to, so_thua=so_thua,
                    gcn_code=None
                )
                note = "Bỏ đơn do có dữ liệu"

            return success, note

    except NoSuchWindowException:
        error_message = "Lỗi: Cửa sổ trình duyệt đã bị đóng đột ngột."
        logger.log(f"❌ {error_message} (Thửa {so_thua}, Tờ {so_to})")
        return True, "Lỗi (Cửa sổ đóng)"
    except Exception as ex:
        logger.log(f"❌ Có lỗi xảy ra khi xử lý thửa {so_thua}, tờ {so_to}: {ex}")
        logger.log(traceback.format_exc())
        return True, f"Lỗi khi xử lý thửa tờ {so_to}, thửa {so_thua}"

# ============== TKINTER UI ==============
def main():
    root = tk.Tk()
    root.title("Tự động xóa đơn - MPLIS")
    root.geometry("800x650")

    excel_file_path = tk.StringVar()

    main_frm = ttk.Frame(root, padding=12)
    main_frm.pack(fill="both", expand=True)
    main_frm.columnconfigure(1, weight=1)

    ttk.Label(main_frm, text="Username:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
    ent_user = ttk.Entry(main_frm, width=40)
    ent_user.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ent_user.insert(0, "")

    ttk.Label(main_frm, text="Password:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
    ent_pass = ttk.Entry(main_frm, width=40, show="•")
    ent_pass.grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ent_pass.insert(0, "")

    ttk.Label(main_frm, text="Tỉnh/Thành phố:").grid(row=2, column=0, sticky="e", padx=4, pady=4)
    province_cb = ttk.Combobox(main_frm, state="readonly", width=37)
    province_cb["values"] = ["Đắk Lắk", "Phú Yên"]
    province_cb.grid(row=2, column=1, sticky="w", padx=4, pady=4)
    province_cb.set("Đắk Lắk")

    ttk.Label(main_frm, text="Mã xã:").grid(row=3, column=0, sticky="e", padx=4, pady=4)
    ent_ma_xa = ttk.Entry(main_frm, width=40)
    ent_ma_xa.grid(row=3, column=1, sticky="w", padx=4, pady=4)

    excel_frm = ttk.LabelFrame(main_frm, text="Cấu hình Excel", padding=10)
    excel_frm.grid(row=4, column=0, columnspan=2, sticky="ew", padx=4, pady=10)
    excel_frm.columnconfigure(1, weight=1)

    btn_browse = ttk.Button(excel_frm, text="Chọn file Excel")
    btn_browse.grid(row=0, column=0, padx=4, pady=4)
    lbl_file_path = ttk.Label(excel_frm, textvariable=excel_file_path, relief="sunken", padding=2)
    lbl_file_path.grid(row=0, column=1, columnspan=3, sticky="ew", padx=4, pady=4)

    ttk.Label(excel_frm, text="Tên cột Số tờ:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
    ent_col_so_to = ttk.Entry(excel_frm, width=20)
    ent_col_so_to.grid(row=1, column=1, sticky="w", padx=4, pady=4)
    ent_col_so_to.insert(0, "soto")

    ttk.Label(excel_frm, text="Tên cột Số thửa:").grid(row=1, column=2, sticky="e", padx=4, pady=4)
    ent_col_so_thua = ttk.Entry(excel_frm, width=20)
    ent_col_so_thua.grid(row=1, column=3, sticky="w", padx=4, pady=4)
    ent_col_so_thua.insert(0, "sothua")

    btn_run = ttk.Button(main_frm, text="Chạy tự động")
    btn_run.grid(row=5, column=1, sticky="w", padx=4, pady=8)

    log_frm = ttk.Frame(main_frm)
    log_frm.grid(row=6, column=0, columnspan=2, sticky="nsew")
    log_frm.columnconfigure(0, weight=1)
    log_frm.rowconfigure(0, weight=1)
    main_frm.rowconfigure(6, weight=1)

    txt = tk.Text(log_frm, state="disabled", bg="#0f1115", fg="#e5e7eb", height=15)
    txt.grid(row=0, column=0, sticky="nsew")
    
    scrollbar = ttk.Scrollbar(log_frm, orient="vertical", command=txt.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    txt['yscrollcommand'] = scrollbar.set

    log = UILogger(txt)

    def select_excel_file():
        if openpyxl is None:
            messagebox.showerror("Thiếu thư viện",
                                 "Vui lòng cài đặt thư viện 'openpyxl' để có thể xử lý file Excel.\n"
                                 "Chạy lệnh sau trong terminal:\n"
                                 "pip install openpyxl")
            return
        filepath = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if filepath:
            excel_file_path.set(filepath)

    def on_run():
        username = ent_user.get().strip()
        password = ent_pass.get()
        province = province_cb.get()
        ma_xa = ent_ma_xa.get().strip()
        file_path = excel_file_path.get()
        col_so_to_orig = ent_col_so_to.get().strip()
        col_so_thua_orig = ent_col_so_thua.get().strip()

        if not all([username, password, province, ma_xa]):
            messagebox.showerror("Thiếu thông tin", "Vui lòng nhập đủ Username, Password, Tỉnh và Mã xã.")
            return
        if not file_path or not col_so_to_orig or not col_so_thua_orig:
            messagebox.showerror("Thiếu thông tin Excel", "Vui lòng chọn file Excel và nhập tên các cột.")
            return

        col_so_to = col_so_to_orig.lower()
        col_so_thua = col_so_thua_orig.lower()

        if province == "Phú Yên":
            base_url = "https://phy.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"
        else:
            base_url = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"

        btn_run.config(state="disabled")
        log.log(f"=== BẮT ĐẦU CHẠY ({province} - Mã xã: {ma_xa}) ===")

        def runner():
            driver = None
            try:
                log.log(f"Đang đọc file: {file_path}")
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                headers = [str(cell.value).lower() if cell.value is not None else '' for cell in sheet[1]]

                if col_so_to not in headers or col_so_thua not in headers:
                    log.log(f"Lỗi: Không tìm thấy cột '{col_so_to_orig}' hoặc '{col_so_thua_orig}' trong file Excel.")
                    log.log(f"Các cột có trong file (đã chuyển thành chữ thường): {headers}")
                    btn_run.config(state="normal")
                    return

                idx_so_to = headers.index(col_so_to)
                idx_so_thua = headers.index(col_so_thua)

                plots_to_process = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    if row_is_highlighted(row):
                        continue

                    so_to_val = row[idx_so_to].value
                    so_thua_val = row[idx_so_thua].value
                    if so_to_val and so_thua_val:
                        plots_to_process.append((row_idx, str(so_to_val).strip(), str(so_thua_val).strip()))

                log.log(f"Tìm thấy {len(plots_to_process)} thửa đất để xử lý.")
                if not plots_to_process:
                    btn_run.config(state="normal")
                    return

                file_name_only = os.path.basename(file_path)
                file_root, file_ext = os.path.splitext(file_name_only)

                result_path = os.path.join(
                    os.path.dirname(file_path),
                    f"{ma_xa}_{file_root}.xlsx"
                )

                log.log(f"📄 File kết quả: {result_path}")

                # --- Chuẩn bị workbook KẾT QUẢ ---
                next_stt = 1

                if os.path.exists(result_path):
                    log.log("📄 File kết quả đã tồn tại, sẽ ghi nối tiếp vào cuối file.")
                    result_wb = openpyxl.load_workbook(result_path)
                    result_ws = result_wb.active

                    last_row = result_ws.max_row
                    last_stt = result_ws.cell(row=last_row, column=1).value

                    if isinstance(last_stt, int):
                        next_stt = last_stt + 1
                    else:
                        next_stt = last_row
                else:
                    result_wb = openpyxl.Workbook()
                    result_ws = result_wb.active
                    result_ws.title = "Ket_qua"
                    result_ws.append(["STT", "Dòng Excel", "Số tờ", "Số thửa", "Ghi chú"])
                    log.log("📄 Chưa có file kết quả, tạo file mới.")

                log.log("🚀 Khởi động Chrome…")
                options = Options()
                options.add_argument("--start-maximized")
                options.add_experimental_option("detach", True)
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=options)
                wait = WebDriverWait(driver, 20)

                driver.get(base_url)
                log.log(f"🌐 Mở trang: {base_url}")

                wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(username)
                driver.find_element(By.NAME, "password").send_keys(password)
                driver.find_element(By.NAME, "password").send_keys(Keys.ENTER)
                log.log("🔐 Đang đăng nhập…")
                messagebox.showinfo("Xác minh",
                                    "Nếu có xác minh thủ công (captcha/SSO), hãy hoàn tất trên trình duyệt rồi bấm OK để tiếp tục.")

                log.log(f"✅ Đăng nhập thành công. Bắt đầu chọn xã có mã: {ma_xa}")
                option_xpath = f"//select[@id='ddlPhuongXaKeKhai']/option[@value='{ma_xa}']"
                option_element = wait.until(EC.element_to_be_clickable((By.XPATH, option_xpath)))
                option_element.click()
                log.log(f"✅ Đã chọn xã có mã: {ma_xa}.")

                log.log("🔎 Mở cửa sổ tra cứu…")
                tra_cuu_button = wait.until(EC.element_to_be_clickable((By.ID, "btnChonDonDangKy")))
                try:
                    tra_cuu_button.click()
                except ElementClickInterceptedException:
                    log.log("⚠️ Click bị chặn, thử lại bằng JavaScript...")
                    driver.execute_script("arguments[0].click();", tra_cuu_button)
                wait_tracuu_module_ready(driver, timeout=60)

                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

                for i, (row_num, so_to, so_thua) in enumerate(plots_to_process):
                    log.log(f"--- Xử lý thửa {i+1}/{len(plots_to_process)}: Tờ {so_to}, Thửa {so_thua} (Dòng {row_num}) ---")

                    processed, note = search_and_process_plot(driver, wait, log, so_to, so_thua)
                    log.log(f"📌 Ghi chú kết quả: {note}")

                    # Ghi kết quả dùng next_stt (kể cả chạy lại nhiều lần)
                    result_ws.append([next_stt, row_num, so_to, so_thua, note])
                    next_stt += 1

                    if not str(note).lower().startswith("lỗi"):
                        log.log(f"🎨 Tô màu dòng {row_num} trong file Excel.")
                        for cell in sheet[row_num]:
                            cell.fill = yellow_fill
                    else:
                        log.log(f"⚠️ Dòng {row_num} có lỗi, KHÔNG tô màu để xử lý lại sau.")

                    if (i + 1) % 50 == 0:
                        try:
                            workbook.save(file_path)
                            log.log(f"💾 Đã lưu file gốc sau khi xử lý {i+1} dòng.")
                        except Exception as save_err:
                            log.log(f"⚠️ Lỗi khi lưu file Excel gốc: {save_err}")

                        try:
                            result_wb.save(result_path)
                            log.log(f"💾 Đã lưu file kết quả sau {i+1} thửa: {result_path}")
                        except Exception as save_err:
                            log.log(f"⚠️ Lỗi khi lưu file Excel kết quả: {save_err}")

                    if processed:
                        log.log("🔄 Mở lại cửa sổ tra cứu cho thửa tiếp theo...")
                        tra_cuu_button = wait.until(EC.element_to_be_clickable((By.ID, "btnChonDonDangKy")))
                        try:
                            tra_cuu_button.click()
                        except ElementClickInterceptedException:
                            log.log("⚠️ Click bị chặn khi mở lại, thử lại bằng JavaScript...")
                            driver.execute_script("arguments[0].click();", tra_cuu_button)
                        wait_tracuu_module_ready(driver, timeout=60)

                try:
                    workbook.save(file_path)
                    log.log("✅ Đã lưu file Excel gốc lần cuối sau khi hoàn tất toàn bộ.")
                except Exception as save_err:
                    log.log(f"⚠️ Lỗi khi lưu file Excel gốc lần cuối: {save_err}")

                try:
                    result_wb.save(result_path)
                    log.log(f"✅ Đã lưu file Excel KẾT QUẢ lần cuối: {result_path}")
                except Exception as save_err:
                    log.log(f"⚠️ Lỗi khi lưu file Excel KẾT QUẢ lần cuối: {save_err}")
                
                log.log("✅✅✅ HOÀN TẤT TOÀN BỘ QUÁ TRÌNH! ✅✅✅")

            except Exception as e:
                log.log(f"❌ Lỗi nghiêm trọng trong quá trình chạy: {e}")
                log.log(traceback.format_exc())
            finally:
                if driver:
                    log.log("Trình duyệt vẫn mở. Đóng chương trình để thoát.")
                btn_run.config(state="normal")

        threading.Thread(target=runner, daemon=True).start()

    btn_browse.configure(command=select_excel_file)
    btn_run.configure(command=on_run)
    root.mainloop()

if __name__ == "__main__":
    main()
