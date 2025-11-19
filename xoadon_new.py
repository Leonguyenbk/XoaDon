import time, traceback, threading, sys, json, re, os
from venv import logger
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, ElementClickInterceptedException, JavascriptException,
    StaleElementReferenceException, NoSuchElementException, ElementNotInteractableException
)

# ---- Tkinter GUI ----
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# ---- Excel Import ----
try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    openpyxl = None

# ============== LOG UI HELPERS ==============
class UILogger:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def log(self, msg):
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.encode(sys.stdout.encoding, errors='replace').decode(sys.stdout.encoding))
        if self.text_widget:
            self.text_widget.after(0, lambda: self._append(msg))

    def _append(self, msg):
        self.text_widget.configure(state="normal")
        self.text_widget.insert("end", msg + "\n")
        self.text_widget.see("end")
        self.text_widget.configure(state="disabled")

# ============== WAITERS / HELPERS ==============
def wait_xuly_modal(driver, timeout=20):
    """
    Đợi modal Xử lý đơn đăng ký hiển thị; trả về WebElement modal.
    Modal có id động bắt đầu bằng 'mdlXuLyDonDangKy-'.
    """
    wait = WebDriverWait(driver, timeout)
    driver.switch_to.default_content()
    modal = wait.until(EC.visibility_of_element_located((
        By.CSS_SELECTOR, "div.modal.modal-fullscreen.in[id^='mdlXuLyDonDangKy-'][style*='display: block']"
    )))
    # đảm bảo body không còn overlay che click
    try:
        WebDriverWait(driver, 5).until(lambda d: d.execute_script("return (window.jQuery? jQuery.active:0)") == 0)
    except Exception:
        pass
    return modal

def wait_jstree_ready_in(container_el, timeout=20):
    """
    Đợi #treeDonDangKy trong container có ít nhất một anchor khác 'Không có dữ liệu'.
    """
    end = time.time() + timeout
    while time.time() < end:
        trees = container_el.find_elements(By.CSS_SELECTOR, "#treeDonDangKy")
        if trees:
            anchors = trees[0].find_elements(By.CSS_SELECTOR, "a.jstree-anchor")
            if anchors:
                if not (len(anchors) == 1 and "Không có dữ liệu" in (anchors[0].text or "")):
                    return trees[0]
        time.sleep(0.2)
    raise TimeoutException("jsTree chưa có dữ liệu trong thời gian cho phép.")

def find_tt_dangky_anchor(tree_el):
    """
    Trả về <a> node 'Thông tin đăng ký' (trong đó text ở <b> bên trong).
    Linh hoạt với phần tử phụ như <div id='elementStatus'>.
    """
    xpaths = [
        ".//a[.//b[normalize-space()='Thông tin đăng ký']]",                     # case phổ biến
        ".//a[normalize-space()='Thông tin đăng ký']",                           # đôi khi text flatten
        ".//a[contains(normalize-space(.), 'Thông tin đăng ký')]",               # lỏng
    ]
    for xp in xpaths:
        els = tree_el.find_elements(By.XPATH, xp)
        if els:
            return els[0]
    raise NoSuchElementException("Không tìm thấy anchor 'Thông tin đăng ký' trong jsTree.")

def wait_page_idle(driver, wait, extra_ms=300):
    wait.until(lambda x: x.execute_script("return document.readyState") == "complete")
    time.sleep(extra_ms/1000.0)

def switch_to_iframe_containing_table(driver, table_id="tblTTThuaDat", timeout=10):
    # quay về top trước
    driver.switch_to.default_content()
    iframes = driver.find_elements(By.TAG_NAME, "iframe")
    deadline = time.time() + timeout
    for idx in range(len(iframes)):
        if time.time() > deadline:
            break
        driver.switch_to.default_content()
        iframes = driver.find_elements(By.TAG_NAME, "iframe")  # refresh
        try:
            driver.switch_to.frame(iframes[idx])
            # kiểm tra có bảng không
            if driver.find_elements(By.CSS_SELECTOR, f"#{table_id}"):
                return True
            # nếu còn iframe lồng nhau
            inner_iframes = driver.find_elements(By.TAG_NAME, "iframe")
            for j in range(len(inner_iframes)):
                driver.switch_to.frame(inner_iframes[j])
                if driver.find_elements(By.CSS_SELECTOR, f"#{table_id}"):
                    return True
                driver.switch_to.parent_frame()
        except Exception:
            continue
    driver.switch_to.default_content()
    return False

def wait_for_table_loaded(driver, table_id="tblTTThuaDat", timeout=15):
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, f"#{table_id}_processing"))
        )
    except TimeoutException:
        pass

def safe_click_row_css(driver, wait, row_css="#tblTraCuuDotBanGiao tbody tr", logger=None):
    wait_page_idle(driver, wait, 300)
    row = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, row_css)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row)
    cell = row.find_element(By.CSS_SELECTOR, "td:nth-child(2)")
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//table[@id='tblTraCuuDotBanGiao']//tbody//tr[1]//td[2]")))
        cell.click()
        return
    except ElementClickInterceptedException:
        wait_page_idle(driver, wait, 300)
        try:
            cell.click()
            return
        except ElementClickInterceptedException:
            pass
    try:
        driver.execute_script("""
            document.querySelectorAll('.jquery-loading-modal__bg')
                  .forEach(el => { el.style.pointerEvents='none'; el.style.display='none'; });
        """)
    except JavascriptException:
        pass
    try:
        driver.execute_script("arguments[0].click();", cell)
        return
    except Exception:
        pass
    first_cell = row.find_element(By.CSS_SELECTOR, "td:nth-child(1)")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", first_cell)
    driver.execute_script("arguments[0].click();", first_cell)

def goto_page(driver, page_number, table_id="tblTTThuaDat", verify_timeout=5):
    driver.execute_script(f"""
        if (window.jQuery && jQuery.fn.DataTable) {{
            var table = jQuery('#{table_id}').DataTable();
            var info  = table.page.info();
            var maxp  = info.pages || 1;
            var target = Math.max(0, Math.min({page_number}-1, maxp-1));
            table.page(target).draw('page');
        }}
    """)
    # verify page changed
    end = time.time() + verify_timeout
    target0 = max(0, page_number-1)
    while time.time() < end:
        ok = driver.execute_script(f"""
            try {{
                var t = jQuery('#{table_id}').DataTable();
                return t.page.info().page;
            }} catch(e){{ return -1; }}
        """)
        if ok == target0:
            return True
        time.sleep(0.2)
    return False

def go_next_datatables(driver, table_id="tblTTThuaDat", timeout=15):
    wait = WebDriverWait(driver, timeout)
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, f"#{table_id}_processing")))
    except TimeoutException:
        pass
    li_next = wait.until(EC.presence_of_element_located((By.ID, f"{table_id}_next")))
    if "disabled" in (li_next.get_attribute("class") or ""):
        return False
    a_next = li_next.find_element(By.TAG_NAME, "a")
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"#{table_id}_next a")))
    w, h, vis = driver.execute_script("""
        const a = arguments[0];
        const r = a.getBoundingClientRect();
        const style = window.getComputedStyle(a);
        return [r.width, r.height, style.visibility !== 'hidden' && style.display !== 'none'];
    """, a_next)
    if not (w > 0 and h > 0 and vis):
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", a_next)
        wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"#{table_id}_next a")))
    first_row = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"#{table_id} tbody tr")))
    try:
        a_next.click()
    except Exception:
        driver.execute_script("arguments[0].click();", a_next)
    try:
        wait.until(EC.staleness_of(first_row))
    except (TimeoutException, StaleElementReferenceException):
        try:
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, f"#{table_id}_processing")))
        except TimeoutException:
            pass
    return True

def handle_whole_page_action(driver, logger: UILogger, table_id="tblTTThuaDat", timeout=15):
    """
    Chọn tất cả các hàng trên trang hiện tại (Shift+Click), sau đó lặp qua
    và bỏ chọn (Ctrl+Click) những hàng đã có trạng thái "Đã duyệt ghi đè"
    để chỉ giữ lại các hàng "Chưa xử lý".
    """
    wait = WebDriverWait(driver, timeout)
    wait.until(EC.presence_of_element_located((By.ID, table_id)))
    rows = driver.find_elements(By.CSS_SELECTOR, f"#{table_id} tbody > tr:not(.child)")

    # Lọc các hàng đang hiển thị và có thể tương tác
    visible_rows = []
    for r in rows:
        try:
            tds = r.find_elements(By.CSS_SELECTOR, "td")
            if tds and r.is_displayed():
                visible_rows.append((r, tds))
        except StaleElementReferenceException:
            continue

    if len(visible_rows) < 1:
        logger.log("   (Không có hàng nào hiển thị để chọn)")
        return 0

    first_row, first_tds = visible_rows[0]
    last_row, last_tds = visible_rows[-1]

    def pick_click_target(row, tds):
        # Ưu tiên click vào checkbox hoặc button nếu có, fallback về ô đầu tiên
        for css in ["input[type='checkbox']:not([disabled])", "button", "a"]:
            try:
                el = row.find_element(By.CSS_SELECTOR, css)
                if el.is_displayed(): return el
            except NoSuchElementException: pass
        return tds[0]

    first_target = pick_click_target(first_row, first_tds)
    last_target = pick_click_target(last_row, last_tds)

    def ensure_visible_and_sized(el):
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        WebDriverWait(driver, timeout).until(lambda d: d.execute_script("""
            const r = arguments[0].getBoundingClientRect();
            const s = getComputedStyle(arguments[0]);
            return r.width > 0 && r.height > 0 && s.display!=='none' && s.visibility!=='hidden';
        """, el))

    try:
        ensure_visible_and_sized(first_target)
        first_target.click() # Click hàng đầu
        if len(visible_rows) > 1:
            ensure_visible_and_sized(last_target)
            # Giữ SHIFT và click hàng cuối để chọn tất cả
            ActionChains(driver).key_down(Keys.SHIFT).click(last_target).key_up(Keys.SHIFT).perform()
    except Exception as e:
        logger.log(f"   (Lỗi Shift-Click, thử fallback... Lỗi: {e})")
        # Fallback nếu Shift-Click lỗi: chọn từng cái một
        for row, tds in visible_rows:
            try:
                target = pick_click_target(row, tds)
                ensure_visible_and_sized(target)
                target.click()
            except Exception:
                continue

    logger.log("   → Đã chọn tất cả, bắt đầu lọc bỏ những bản ghi đã duyệt...")
    time.sleep(0.2) # Chờ một chút để UI cập nhật trạng thái "selected"

    # Bỏ chọn những hàng đã được duyệt
    actions = ActionChains(driver).key_down(Keys.CONTROL)
    deselected_count = 0
    # Lấy lại danh sách hàng đã chọn (có class 'selected')
    selected_rows = driver.find_elements(By.CSS_SELECTOR, f"#{table_id} tbody tr.selected")
    for row in selected_rows:
        try:
            txt = (row.get_attribute("innerText") or row.text).strip().lower()
            if "đã duyệt ghi đè" in txt:
                actions.click(row.find_element(By.CSS_SELECTOR, "td:first-child"))
                deselected_count += 1
        except (StaleElementReferenceException, NoSuchElementException):
            continue
    actions.key_up(Keys.CONTROL).perform()

    # Kiểm tra lại số lượng đã chọn bằng API của DataTable
    selected_count = driver.execute_script(f"""
        try {{
            if (window.jQuery && jQuery.fn.DataTable) {{
                const dt = jQuery("#{table_id}").DataTable();
                return dt.rows({{selected:true, page:'current'}}).count();
            }}
        }} catch(e) {{}}
        const table = document.querySelector("#{table_id}");
        return table ? table.querySelectorAll("tbody tr.selected").length : 0;
    """)

    if deselected_count > 0:
        logger.log(f"   → Đã bỏ chọn {deselected_count} bản ghi đã duyệt. Còn lại {selected_count} bản ghi.")

    return selected_count

def quick_confirm_if_present(driver, root_el=None, soft_timeout=1.2):
    """
    Tìm & bấm nút xác nhận nếu có (SweetAlert2/Bootstrap). KHÔNG raise TimeoutException.
    Trả về True nếu đã bấm xác nhận; False nếu không thấy gì để bấm.
    root_el: nếu truyền modal WebElement, chỉ tìm trong đó (ổn định hơn).
    """
    try:
        scope = root_el if root_el is not None else driver
        sw = WebDriverWait(driver, soft_timeout)

        # 1) SweetAlert2 .swal2-confirm
        btns = scope.find_elements(By.CSS_SELECTOR, ".swal2-container .swal2-confirm")
        if not btns:
            # 2) Bootstrap modal primary
            btns = scope.find_elements(By.CSS_SELECTOR, ".modal.in .btn-primary, .modal.show .btn-primary")

        if not btns:
            # 3) Theo text tiếng Việt/English phổ biến
            xp = ".//button[normalize-space()='Đồng ý' or normalize-space()='Xác nhận' or normalize-space()='OK' or normalize-space()='Có' or normalize-space()='Yes']"
            try:
                btns = scope.find_elements(By.XPATH, xp)
            except Exception:
                btns = []

        if not btns:
            # Không thấy gì → coi như không có confirm
            return False

        # Chọn nút hiển thị được
        cand = None
        for b in btns:
            try:
                vis = driver.execute_script("""
                    const el = arguments[0];
                    const r = el.getBoundingClientRect();
                    const s = getComputedStyle(el);
                    return r.width>0 && r.height>0 && s.visibility!=='hidden' && s.display!=='none';
                """, b)
                if vis:
                    cand = b
                    break
            except Exception:
                continue
        if cand is None:
            return False

        # Đảm bảo không bị backdrop che
        try:
            driver.execute_script("""
                document.querySelectorAll('.modal-backdrop, .swal2-container, .jquery-loading-modal__bg')
                    .forEach(el=>{ el.style.pointerEvents='auto'; });
            """)
        except Exception:
            pass

        # Thử click thường
        try:
            cand.click()
            return True
        except Exception:
            pass

        # Thử JS click
        try:
            driver.execute_script("arguments[0].click();", cand)
            return True
        except Exception:
            pass

        # Thử phím Enter vào phần tử đang focus/active
        try:
            driver.switch_to.active_element.send_keys(Keys.ENTER)
            return True
        except Exception:
            pass

        return False
    except Exception:
        # Tuyệt đối không để propagate TimeoutException từ waits bên trong
        return False

def wait_processing_quick(driver, table_id="tblTTThuaDat", max_wait=6):
    def cond(d):
        try:
            ajax_zero = d.execute_script("return (window.jQuery ? jQuery.active : 0)") == 0
            proc = d.execute_script(f"""
                var e = document.querySelector('#{table_id}_processing');
                if(!e) return true;
                var s = getComputedStyle(e);
                return (s.display==='none' || s.visibility==='hidden' || e.offsetParent===null);
            """)
            return ajax_zero and proc
        except Exception:
            return True
    try:
        WebDriverWait(driver, max_wait, poll_frequency=0.1).until(cond)
        return True
    except Exception:
        return False

def hard_jump_pagination(driver, page_number, table_id="tblTTThuaDat", timeout=10):
    wait = WebDriverWait(driver, timeout)
    # xác định trang hiện tại
    cur = driver.execute_script(f"""
        try {{
            return jQuery('#{table_id}').DataTable().page.info().page + 1;
        }} catch(e) {{ return 1; }}
    """) or 1

    if page_number == cur:
        return True

    # nếu có nút số trang, thử click trực tiếp
    try:
        btn = wait.until(EC.presence_of_element_located((
            By.XPATH, f"//div[@id='{table_id}_paginate']//a[normalize-space(text())='{page_number}']"
        )))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        try:
            btn.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn)
    except TimeoutException:
        # nếu không có nút số trang (hiển thị dạng next/prev) → lặp next/prev
        step = 1 if page_number > cur else -1
        next_sel = f"#{table_id}_next a"
        prev_sel = f"#{table_id}_previous a"
        while cur != page_number:
            sel = next_sel if step == 1 else prev_sel
            try:
                a = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
                a.click()
            except Exception:
                driver.execute_script("document.querySelector(arguments[0])?.click()", sel)
            wait_for_table_loaded(driver, table_id, timeout=10)
            cur = driver.execute_script(f"return jQuery('#{table_id}').DataTable().page.info().page + 1;") or cur
            # tránh lặp vô hạn
            if (step == 1 and cur < page_number) or (step == -1 and cur > page_number):
                continue
            if cur == page_number:
                break

    wait_for_table_loaded(driver, table_id, timeout=10)
    cur2 = driver.execute_script(f"return jQuery('#{table_id}').DataTable().page.info().page + 1;")
    return cur2 == page_number

def all_jconfirm_closed(driver):
    """True nếu không còn popup jConfirm nào đang hiển thị."""
    modals = driver.find_elements(By.CSS_SELECTOR, ".jconfirm-scrollpane")
    if not modals:
        return True
    for m in modals:
        try:
            if m.is_displayed():
                return False
        except Exception:
            continue
    return True

def wait_all_jconfirm_closed(driver, timeout=15):
    """
    Chờ cho đến khi KHÔNG còn popup jConfirm nào hiển thị.
    Dùng wrapper cho all_jconfirm_closed để đảm bảo không còn overlay che click.
    """
    try:
        WebDriverWait(driver, timeout).until(lambda d: all_jconfirm_closed(d))
    except TimeoutException:
        # Hết thời gian vẫn còn popup thì thôi, không raise để khỏi vỡ flow
        pass

def switch_to_frame_having(driver, by, value, timeout=8):
    driver.switch_to.default_content()
    # thử ở top trước
    try:
        if driver.find_elements(by, value):
            return True
    except Exception:
        pass
    # duyệt qua tất cả iframes (kể cả lồng nhau)
    frames = driver.find_elements(By.TAG_NAME, "iframe")
    deadline = time.time() + timeout
    for i in range(len(frames)):
        if time.time() > deadline: break
        driver.switch_to.default_content()
        frames = driver.find_elements(By.TAG_NAME, "iframe")  # refresh
        try:
            driver.switch_to.frame(frames[i])
            if driver.find_elements(by, value):
                return True
            # thử thêm 1 tầng lồng
            inner = driver.find_elements(By.TAG_NAME, "iframe")
            for j in range(len(inner)):
                driver.switch_to.frame(inner[j])
                if driver.find_elements(by, value):
                    return True
                driver.switch_to.parent_frame()
        except Exception:
            continue
    driver.switch_to.default_content()
    return False

def wait_tracuu_module_ready(driver, timeout=60):
    # Chờ phần tử xuất hiện trong DOM
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#donDangKyTraCuuModule"))
    )

    # Chờ nó thực sự visible
    WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#donDangKyTraCuuModule"))
    )

    # Chờ không còn overlay loading
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".jquery-loading-modal__bg"))
        )
    except:
        pass

    # Chờ module render xong (DOM height > 0)
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let el = document.querySelector("#donDangKyTraCuuModule");
            if (!el) return false;
            return el.offsetHeight > 0 && el.offsetWidth > 0;
        """)
    )

    print("✅ Module tra cứu (#donDangKyTraCuuModule) đã load xong!")

def wait_tracuu_section_ready(driver, timeout=60):
    selector = "#donDangKyTraCuuModule > div.panel-body > div > div:nth-child(3)"

    # 1) Chờ xuất hiện trong DOM
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
    )

    # 2) Chờ nó visible thật sự
    WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, selector))
    )

    # 3) Chờ overlay biến mất (nếu có)
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".jquery-loading-modal__bg"))
        )
    except:
        pass

    # 4) Chờ height/width > 0 (DOM render xong)
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let el = document.querySelector(arguments[0]);
            if (!el) return false;
            let rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        """, selector)
    )

    print("✅ Vùng tra cứu (div:nth-child(3)) đã load xong!")

def wait_and_count_tblTraCuu(driver, timeout=60):
    table_selector = "#tblTraCuuTinhHinhDangKy"

    # 1) Chờ bảng xuất hiện
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, table_selector))
    )

    # 2) Chờ overlay MPLIS biến mất
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((By.CSS_SELECTOR, ".jquery-loading-modal__bg"))
        )
    except:
        pass

    # 3) Chờ DataTables ngừng processing
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let p = document.querySelector("#tblTraCuuTinhHinhDangKy_processing");
            if (p && p.offsetParent !== null) return false;  // đang loading
            return true;
        """)
    )

    # 4) Chờ tbody xuất hiện
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("""
            let tb = document.querySelector("#tblTraCuuTinhHinhDangKy tbody");
            return tb && tb.children.length >= 0;
        """)
    )

    # 5) Đếm số bản ghi thật
    count = driver.execute_script("""
        let table = document.querySelector("#tblTraCuuTinhHinhDangKy");
        if (!table) return -1;

        let rows = table.querySelectorAll("tbody tr");
        if (!rows || rows.length === 0) return 0;

        let count = 0;
        rows.forEach(r => {
            let td = r.querySelector("td");
            if (td && td.classList.contains("dataTables_empty")) return; 
            count++;
        });

        return count;
    """)

    print("➡️ Số bản ghi:", count)
    return count

def wait_query_done(driver, timeout=30, ajax_wait=5):
    """
    Chờ các request AJAX (jQuery) phục vụ việc tra cứu đơn hoàn tất.
    - Gọi NGAY SAU khi click nút Tra cứu.
    - Không phụ thuộc vào việc bảng có thay nội dung hay không.
    - Không đụng tới DataTables API nên tránh lỗi _DT_CellIndex.
    """

    end_time = time.time() + timeout

    # 1. Đợi jQuery có trên trang (nếu vì lý do gì đó chưa load)
    try:
        WebDriverWait(driver, 5).until(
            lambda d: d.execute_script("return window.jQuery !== undefined;")
        )
    except Exception:
        # Không có jQuery thì coi như không chờ được AJAX, thoát nhẹ nhàng
        return

    # 2. Pha 1: cố gắng đợi có ÍT NHẤT 1 request AJAX bắt đầu (jQuery.active > 0)
    #    nhưng tối đa ajax_wait giây, nếu không thấy thì thôi, coi như không có AJAX
    phase1_end = time.time() + ajax_wait
    saw_ajax = False
    while time.time() < phase1_end:
        try:
            active = driver.execute_script("return jQuery.active;")
            if active > 0:
                saw_ajax = True
                break
        except Exception:
            # jQuery biến mất hay gì đó, thôi không chờ nữa
            break
        time.sleep(0.1)

    if not saw_ajax:
        # Không thấy request nào bắt đầu trong ajax_wait giây -> có thể kết quả được cache
        # hoặc xử lý đồng bộ, ta không chờ nữa.
        return

    # 3. Pha 2: đã thấy AJAX bắt đầu -> giờ đợi đến khi tất cả AJAX xong (jQuery.active == 0)
    while time.time() < end_time:
        try:
            active = driver.execute_script("return jQuery.active;")
            if active == 0:
                return
        except Exception:
            # Nếu jQuery không còn, nhiều khả năng trang đã xong / chuyển trang
            return
        time.sleep(0.1)

def wait_query_xoadon(driver, timeout=30, ajax_wait=5, max_after_first_ajax=10):
    """
    Chờ các request AJAX (jQuery) phục vụ việc tra cứu đơn hoàn tất.

    - Gọi NGAY SAU khi click nút Tra cứu.
    - Pha 1: đợi phát hiện ÍT NHẤT 1 request AJAX bắt đầu (jQuery.active > 0)
             trong tối đa ajax_wait giây.
    - Pha 2: sau khi thấy AJAX bắt đầu, đợi tối đa max_after_first_ajax giây
             để jQuery.active giảm xuống (0 hoặc gần 0) rồi thoát.
    - Tổng thời gian sẽ bị khống chế bởi (ajax_wait + max_after_first_ajax),
      KHÔNG bao giờ kéo dài hết timeout như trước.
    """

    # 1. Đợi jQuery có trên trang
    try:
        WebDriverWait(driver, 3).until(
            lambda d: d.execute_script("return window.jQuery !== undefined;")
        )
    except Exception:
        # Không có jQuery thì coi như không chờ được AJAX
        return

    # 2. Pha 1: cố gắng đợi có ÍT NHẤT 1 request AJAX bắt đầu
    phase1_end = time.time() + ajax_wait
    saw_ajax = False

    while time.time() < phase1_end:
        try:
            active = driver.execute_script("return jQuery.active;")
            if active > 0:
                saw_ajax = True
                break
        except Exception:
            # jQuery biến mất -> thôi, không chờ nữa
            return
        time.sleep(0.1)

    if not saw_ajax:
        # Không thấy request nào bắt đầu trong ajax_wait giây
        # -> Có thể trang xử lý đồng bộ hoặc cache -> không chờ nữa
        return

    # 3. Pha 2: Đã thấy AJAX bắt đầu -> chờ đến khi nó "lặng" bớt
    #    nhưng TỐI ĐA max_after_first_ajax giây, không đợi hết 30s
    phase2_end = time.time() + max_after_first_ajax

    # Có thể cho phép 1–2 request nền vẫn chạy, nên dùng ngưỡng <= 1
    THRESH = 1
    stable_required = 5   # cần liên tiếp 5 lần (0.1s * 5 = 0.5s) dưới ngưỡng
    stable_count = 0

    while time.time() < phase2_end:
        try:
            active = driver.execute_script("return jQuery.active;")
        except Exception:
            # jQuery không còn -> nhiều khả năng trang xong/chuyển
            return

        if active <= THRESH:
            stable_count += 1
            if stable_count >= stable_required:
                # Đã yên ổn một lúc -> coi như xong
                return
        else:
            # Lại có request mới -> reset bộ đếm
            stable_count = 0

        time.sleep(0.1)

    # Hết max_after_first_ajax giây mà vẫn chưa "yên" hẳn -> kệ, thoát.
    return


def chon_ban_ghi_dau_tien(driver, timeout=30):
    wait = WebDriverWait(driver, timeout)

    # 1. Chờ có ít nhất 1 dòng trong bảng
    first_row = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr")
        )
    )

    # Trường hợp không có bản ghi nào
    if "Không tìm thấy" in first_row.text:
        return False

    # 2. Tìm ô checkbox
    checkbox = wait.until(
        EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr:nth-child(1) td.select-checkbox")
        )
    )

    checkbox.click()

    # 3. Chờ DataTables thêm class 'selected'
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "#tblTraCuuTinhHinhDangKy tbody tr.selected")
        )
    )

    # 4. Nhấn nút "Chọn"
    btn_chon = wait.until(
        EC.element_to_be_clickable((By.ID, "btnLuuChonTinhHinhDangKy"))
    )
    btn_chon.click()

    # 5. Chờ modal đóng (panel ẩn đi)
    wait.until(
        EC.invisibility_of_element_located((By.ID, "donDangKyTraCuuModule"))
    )

    return True

def click_step_GiayChungNhan(driver, timeout=30):
    wait = WebDriverWait(driver, timeout)

    selector = "li.tour_kekhaidangky_step16"

    # 1. Đợi step xuất hiện
    step = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))

    # 2. Nếu đã active thì thôi
    if "active" in step.get_attribute("class"):
        return True

    # 3. Click
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector))).click()

    # 4. Chờ trở thành active
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, selector + ".clickable.active")
        )
    )

    return True

def kiem_tra_tree_gcn(driver):
    """
    Kiểm tra cây #treeGiayChungNhan.

    Trả về (status, gcn_code, raw_text):
      - status = "no_data"    : text chứa "Không có dữ liệu"/"Không có giữ liệu" => XÓA ĐƠN
      - status = "has_gcn"    : Có chuỗi 'Số phát hành: ...' => BỎ ĐƠN do có mã GCN
      - status = "has_data"   : Có dữ liệu khác (không chứa 'Không có dữ liệu' và không match regex) => BỎ ĐƠN do có dữ liệu
    """
    anchors = driver.find_elements(By.CSS_SELECTOR, "#treeGiayChungNhan a.jstree-anchor")

    if not anchors:
        print("❌ Không có dữ liệu trong #treeGiayChungNhan")
        return ("no_data", None, "")

    text = anchors[0].text.strip()
    text_lower = text.lower()

    # Trường hợp UI ghi 'Không có dữ liệu' (hoặc gõ nhầm 'giữ liệu')
    if "không có dữ liệu" in text_lower or "không có giữ liệu" in text_lower:
        print("ℹ️ Cây GCN hiển thị 'Không có dữ liệu'")
        return ("no_data", None, text)

    # Regex tìm số phát hành
    pattern = r"Số phát hành:\s*((?:[A-Za-zĐđ]{1,2}\s?\d{5,8})|(?:\d{5,8}))"
    match = re.search(pattern, text)

    if match:
        gcn_code = match.group(1).strip()
        print(f"✅ Có dữ liệu GCN, Số phát hành: {gcn_code}")
        return ("has_gcn", gcn_code, text)
    else:
        print("✅ Có dữ liệu trong cây GCN nhưng không tìm thấy 'Số phát hành'")
        return ("has_data", None, text)

def perform_bo_don(driver, wait, logger: UILogger, reason="", so_to=None, so_thua=None, gcn_code=None):
    """
    Hàm riêng để thực hiện thao tác "Bỏ đơn".
    """
    log_message = f"✅ {reason} Tiến hành bỏ đơn..."
    logger.log(log_message)

    # Nếu lý do là có GCN hoặc bị thế chấp, lưu thông tin thửa đất ra file txt
    if ("GCN" in reason or "thế chấp" in reason) and so_to and so_thua:
        try:
            with open("thua_dat_co_gcn.txt", "a", encoding="utf-8") as f:
                f.write(f"Số tờ: {so_to}, Số thửa: {so_thua}, Mã GCN: {gcn_code or 'N/A'}\n")
            logger.log(f"💾 Đã lưu thông tin thửa đất có GCN vào file 'thua_dat_co_gcn.txt'.")
        except Exception as e:
            logger.log(f"⚠️ Lỗi khi ghi file txt: {e}")
            print(f"⚠️ Lỗi khi ghi file txt: {e}")
    
    try:
        btn_bo_don = wait.until(EC.element_to_be_clickable((By.ID, "btnBoDonDangKy")))
        btn_bo_don.click()

        # 1. Chờ popup xác nhận xuất hiện
        wait.until(
            EC.visibility_of_element_located((
                By.CSS_SELECTOR,
                "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open"
            ))
        )
        print("👉 Popup xác nhận 'Bỏ đơn' đã xuất hiện")

        # 2. Chờ đúng nút cam (btn btn-orange) xuất hiện và có thể click
        btn_orange = wait.until(
            EC.element_to_be_clickable((
                By.CSS_SELECTOR,
                "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open .jconfirm-buttons button.btn.btn-orange"
            ))
        )
        print("👉 Nút cam 'Đồng ý' đã sẵn sàng")

        # 3. Nhấn nút cam
        btn_orange.click()
        print("👉 Đã nhấn nút cam 'Đồng ý'")

        # 4. Chờ popup đóng hoàn toàn
        wait.until(
            EC.invisibility_of_element_located((
                By.CSS_SELECTOR,
                "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open"
            ))
        )
        print("✅ Popup 'Bỏ đơn' đã đóng")

        WebDriverWait(driver, 15).until(lambda d: all_jconfirm_closed(d))
        print("✅ Tất cả popup đã đóng – Bỏ đơn thành công!")
        logger.log("✅ Thao tác 'Bỏ đơn' hoàn tất.")
        return True

    except Exception as e:
        logger.log(f"❌ Lỗi trong quá trình 'Bỏ đơn': {e}")
        print(f"❌ Lỗi trong quá trình 'Bỏ đơn': {e}")
        # vẫn trả về True để vòng lặp chính biết cần mở lại modal
        return True

def search_and_process_plot(driver, wait, logger: UILogger, so_to, so_thua):
    """
    Thực hiện tìm kiếm và xử lý một thửa đất trong modal tra cứu đã mở.

    Trả về:
        processed (bool): True nếu đã có hành động (xóa/bỏ) làm đóng modal.
        note (str): ghi chú để ghi ra file Excel kết quả, gồm:
            - "Không tìm thấy bản ghi"
            - "Bỏ đơn do có dữ liệu"
            - "bỏ đơn do có mã GCN {mã gcn}"
            - "Đã xóa đơn"
            - hoặc "Lỗi khi xử lý thửa ..." (trường hợp ngoại lệ)
    """
    try:
        # --- Nhập liệu và tìm kiếm ---
        so_thua_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,
            "#dvTraCuuTinhHinhDangKyChiTiet > div:nth-child(2) > div.col-md-8.col-sm-12 > fieldset > div:nth-child(2) > div:nth-child(1) > div > input"
        )))
        so_thua_input.clear()
        so_thua_input.send_keys(so_thua)

        so_to_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,
            "#dvTraCuuTinhHinhDangKyChiTiet > div:nth-child(2) > div.col-md-8.col-sm-12 > fieldset > div:nth-child(2) > div:nth-child(2) > div > input"
        )))
        so_to_input.clear()
        so_to_input.send_keys(so_to)

        so_thua_input.send_keys(Keys.ENTER)

        wait_tracuu_section_ready(driver, timeout=60)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "tblTraCuuTinhHinhDangKy_info"))
        )
        wait_query_done(driver, timeout=60)
        so_ban_ghi = wait_and_count_tblTraCuu(driver)
        logger.log(f"✅ Đã nhập Số tờ: {so_to}, Số thửa: {so_thua}. Số bản ghi tìm được: {so_ban_ghi}.")

        if so_ban_ghi == 0:
            logger.log("❌ Không tìm thấy bản ghi nào. Tìm thửa tiếp theo...")
            return False, "Không tìm thấy bản ghi"

        # --- Tìm thấy, xử lý ---
        chon_ban_ghi_dau_tien(driver, timeout=30)
        wait_query_done(driver, timeout=60)
        click_step_GiayChungNhan(driver, timeout=30)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "treeGiayChungNhan"))
        )
        status, gcn_code, raw_text = kiem_tra_tree_gcn(driver)
        if status == "no_data":
            logger.log("👉 Kết quả kiểm tra GCN: Không có dữ liệu (sẽ XÓA ĐƠN).")
        elif status == "has_gcn":
            logger.log(f"👉 Kết quả kiểm tra GCN: Có mã GCN {gcn_code} (sẽ BỎ ĐƠN).")
        else:
            logger.log("👉 Kết quả kiểm tra GCN: Có dữ liệu nhưng không có 'Số phát hành' (sẽ BỎ ĐƠN).")

        # ================== TRƯỜNG HỢP KHÔNG CÓ DỮ LIỆU -> XÓA ĐƠN ==================
        if status == "no_data":
            # ===== XÓA ĐƠN ĐĂNG KÝ =====
            try:
                btn_xoa = wait.until(EC.element_to_be_clickable((By.ID, "btnXoaDonDangKy")))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_xoa)
                btn_xoa.click()
                print("👉 Đã nhấn nút Xóa đơn đăng ký")
            except Exception as e:
                print(f"❌ Không tìm thấy hoặc không click được nút Xóa đơn đăng ký: {e}")
                logger.log("❌ Không tìm thấy nút Xóa đơn đăng ký.")
                # coi như có xử lý nhưng thất bại
                return True, "Lỗi khi xử lý thửa (không click được nút Xóa đơn đăng ký)"
            WebDriverWait(driver, 15).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR,
                    "div.jconfirm.jconfirm-open .jconfirm-scrollpane")))           
            # ---- POPUP 1: ĐỒNG Ý / KHÔNG ----
            try:
                # chờ popup hiện
                wait.until(EC.visibility_of_element_located((
                    By.CSS_SELECTOR, "div.jconfirm.jconfirm-open .jconfirm-scrollpane"
                )))

                dongy_btn = wait.until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "div.jconfirm.jconfirm-open .jconfirm-buttons button.btn.btn-orange"
                )))
                print("👉 Popup xác nhận đã hiện, nhấn ĐỒNG Ý")

                try:
                    dongy_btn.click()
                except ElementClickInterceptedException:
                    driver.execute_script("arguments[0].click();", dongy_btn)                
              
            except Exception as e:
                print(f"❌ Không thấy hoặc không click được nút ĐỒNG Ý: {e}")
                logger.log("❌ Không thấy popup xác nhận khi xóa.")
                return True, "Lỗi khi xử lý thửa (không click được Đồng ý khi xóa)"
            
            wait_query_xoadon(driver, timeout=60)           
            wait_all_jconfirm_closed(driver, timeout=15)

            # ---- POPUP 2: OK ----
            try:
                selector = (
                    "div.jconfirm.jconfirm-vbdlis-theme.jconfirm-open "
                    "div.jconfirm-buttons > button"
                )
                wait = WebDriverWait(driver, 60)
                # Chờ element xuất hiện trong DOM
                btn = wait.until(EC.presence_of_element_located(
                    (By.CSS_SELECTOR, selector)
                ))

                # Chờ nó hiển thị & clickable
                btn = wait.until(EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, selector)
                ))

                btn.click()
                print("👉 Đã nhấn nút OK jConfirm thành công") 

            except Exception as e:
                print(f"❌ Không tìm thấy / không click được nút OK: {e}")
                # vẫn tiếp tục chờ đóng popup phía dưới
            
            wait_all_jconfirm_closed(driver, timeout=15)

            try:
                WebDriverWait(driver, 15).until(lambda d: all_jconfirm_closed(d))
            except Exception:
                print("⚠ Popup không biến mất đúng hạn, nhưng có thể đã xử lý xong")
                logger.log("⚠ Thao tác xóa hoàn tất nhưng popup không tự đóng.")

            # nối theo yêu cầu: text cây GCN 'Không có dữ liệu' => ĐÃ XÓA ĐƠN
            return True, "Đã xóa đơn"

        # ================== TRƯỜNG HỢP CÓ DỮ LIỆU -> BỎ ĐƠN ==================
        else:
            if status == "has_gcn":
                # Thửa đất có GCN cụ thể
                success = perform_bo_don(
                    driver, wait, logger,
                    reason="Thửa đất đã có GCN.",
                    so_to=so_to, so_thua=so_thua,
                    gcn_code=gcn_code
                )
                note = f"bỏ đơn do có mã GCN {gcn_code}"
            else:
                # Có dữ liệu nhưng không có số phát hành
                success = perform_bo_don(
                    driver, wait, logger,
                    reason="Thửa đất có dữ liệu GCN (không có số phát hành).",
                    so_to=so_to, so_thua=so_thua,
                    gcn_code=None
                )
                note = "Bỏ đơn do có dữ liệu"

            return success, note

    except Exception as ex:
        logger.log(f"❌ Có lỗi xảy ra khi xử lý thửa {so_thua}, tờ {so_to}: {ex}")
        logger.log(traceback.format_exc())
        # Coi như đã xử lý (modal có thể đóng), và ghi chú lỗi
        return True, f"Lỗi khi xử lý thửa tờ {so_to}, thửa {so_thua}"

# ============== TKINTER UI ==============
def main():
    root = tk.Tk()
    root.title("Tự động xóa đơn - MPLIS")
    root.geometry("800x650")

    # --- Biến lưu trữ ---
    excel_file_path = tk.StringVar()

    # --- Frame chính ---
    main_frm = ttk.Frame(root, padding=12)
    main_frm.pack(fill="both", expand=True)
    main_frm.columnconfigure(1, weight=1)

    # --- Các trường nhập ---
    ttk.Label(main_frm, text="Username:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
    ent_user = ttk.Entry(main_frm, width=40)
    ent_user.grid(row=0, column=1, sticky="ew", padx=4, pady=4)
    ent_user.insert(0, "")

    ttk.Label(main_frm, text="Password:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
    ent_pass = ttk.Entry(main_frm, width=40, show="•")
    ent_pass.grid(row=1, column=1, sticky="ew", padx=4, pady=4)
    ent_pass.insert(0, "")

    # --- Combobox chọn Tỉnh ---
    ttk.Label(main_frm, text="Tỉnh/Thành phố:").grid(row=2, column=0, sticky="e", padx=4, pady=4)
    province_cb = ttk.Combobox(main_frm, state="readonly", width=37)
    province_cb["values"] = ["Đắk Lắk", "Phú Yên"]
    province_cb.grid(row=2, column=1, sticky="w", padx=4, pady=4)
    province_cb.set("Đắk Lắk") # Mặc định

    # --- Trường nhập Mã xã ---
    ttk.Label(main_frm, text="Mã xã:").grid(row=3, column=0, sticky="e", padx=4, pady=4)
    ent_ma_xa = ttk.Entry(main_frm, width=40)
    ent_ma_xa.grid(row=3, column=1, sticky="w", padx=4, pady=4)

    # --- Frame cấu hình Excel ---
    excel_frm = ttk.LabelFrame(main_frm, text="Cấu hình Excel", padding=10)
    excel_frm.grid(row=4, column=0, columnspan=2, sticky="ew", padx=4, pady=10)
    excel_frm.columnconfigure(1, weight=1)

    btn_browse = ttk.Button(excel_frm, text="Chọn file Excel")
    btn_browse.grid(row=0, column=0, padx=4, pady=4)
    lbl_file_path = ttk.Label(excel_frm, textvariable=excel_file_path, relief="sunken", padding=2)
    lbl_file_path.grid(row=0, column=1, columnspan=3, sticky="ew", padx=4, pady=4)

    ttk.Label(excel_frm, text="Tên cột Số tờ:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
    ent_col_so_to = ttk.Entry(excel_frm, width=20)
    ent_col_so_to.grid(row=1, column=1, sticky="w", padx=4, pady=4)
    ent_col_so_to.insert(0, "soto")

    ttk.Label(excel_frm, text="Tên cột Số thửa:").grid(row=1, column=2, sticky="e", padx=4, pady=4)
    ent_col_so_thua = ttk.Entry(excel_frm, width=20)
    ent_col_so_thua.grid(row=1, column=3, sticky="w", padx=4, pady=4)
    ent_col_so_thua.insert(0, "sothua")

    # --- Nút chạy ---
    btn_run = ttk.Button(main_frm, text="Chạy tự động")
    btn_run.grid(row=5, column=1, sticky="w", padx=4, pady=8)

    # --- Vùng log ---
    log_frm = ttk.Frame(main_frm)
    log_frm.grid(row=6, column=0, columnspan=2, sticky="nsew")
    log_frm.columnconfigure(0, weight=1)
    log_frm.rowconfigure(0, weight=1)
    main_frm.rowconfigure(6, weight=1)

    txt = tk.Text(log_frm, state="disabled", bg="#0f1115", fg="#e5e7eb", height=15)
    txt.grid(row=0, column=0, sticky="nsew")
    
    scrollbar = ttk.Scrollbar(log_frm, orient="vertical", command=txt.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    txt['yscrollcommand'] = scrollbar.set

    logger = UILogger(txt)

    # --- Hàm xử lý ---
    def select_excel_file():
        if openpyxl is None:
            messagebox.showerror("Thiếu thư viện",
                                 "Vui lòng cài đặt thư viện 'openpyxl' để có thể xử lý file Excel.\n"
                                 "Chạy lệnh sau trong terminal:\n"
                                 "pip install openpyxl")
            return
        filepath = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        if filepath:
            excel_file_path.set(filepath)

    def on_run():
        # Lấy thông tin từ GUI
        username = ent_user.get().strip()
        password = ent_pass.get()
        province = province_cb.get()
        ma_xa = ent_ma_xa.get().strip()
        file_path = excel_file_path.get()
        col_so_to_orig = ent_col_so_to.get().strip()
        col_so_thua_orig = ent_col_so_thua.get().strip()

        # Kiểm tra thông tin
        if not all([username, password, province, ma_xa]):
            messagebox.showerror("Thiếu thông tin", "Vui lòng nhập đủ Username, Password, Tỉnh và Mã xã.")
            return
        if not file_path or not col_so_to_orig or not col_so_thua_orig:
            messagebox.showerror("Thiếu thông tin Excel", "Vui lòng chọn file Excel và nhập tên các cột.")
            return

        # Chuyển tên cột sang chữ thường để so sánh không phân biệt hoa thường
        col_so_to = col_so_to_orig.lower()
        col_so_thua = col_so_thua_orig.lower()

        # Chọn URL theo tỉnh
        if province == "Phú Yên":
            base_url = "https://phy.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"
        else: # Mặc định là Đắk Lắk
            base_url = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"

        # Chạy bot trong luồng riêng
        btn_run.config(state="disabled")
        logger.log(f"=== BẮT ĐẦU CHẠY ({province} - Mã xã: {ma_xa}) ===")

        def runner():
            driver = None
            try:
                # --- Đọc dữ liệu Excel ---
                logger.log(f"Đang đọc file: {file_path}")
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                headers = [str(cell.value).lower() if cell.value is not None else '' for cell in sheet[1]]

                if col_so_to not in headers or col_so_thua not in headers:
                    logger.log(f"Lỗi: Không tìm thấy cột '{col_so_to_orig}' hoặc '{col_so_thua_orig}' trong file Excel.")
                    logger.log(f"Các cột có trong file (đã chuyển thành chữ thường): {headers}")
                    btn_run.config(state="normal")
                    return

                idx_so_to = headers.index(col_so_to)
                idx_so_thua = headers.index(col_so_thua)

                plots_to_process = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    so_to_val = row[idx_so_to].value
                    so_thua_val = row[idx_so_thua].value
                    if so_to_val and so_thua_val:
                        plots_to_process.append((row_idx, str(so_to_val).strip(), str(so_thua_val).strip()))

                logger.log(f"Tìm thấy {len(plots_to_process)} thửa đất để xử lý.")
                if not plots_to_process:
                    btn_run.config(state="normal")
                    return

                # --- Chuẩn bị workbook KẾT QUẢ ---
                result_wb = openpyxl.Workbook()
                result_ws = result_wb.active
                result_ws.title = "Ket_qua"
                result_ws.append(["STT", "Dòng Excel", "Số tờ", "Số thửa", "Ghi chú"])

                # Tạo tên file dạng <ma_xa>_<ten_file_goc>.xlsx          
                file_name_only = os.path.basename(file_path)             # ví dụ: danhsach.xlsx
                file_root, file_ext = os.path.splitext(file_name_only)
                
                result_path = os.path.join(
                    os.path.dirname(file_path),
                    f"{ma_xa}_{file_root}.xlsx"                          # ví dụ: 260314_danhsach.xlsx
                )
                logger.log(f"📄 File kết quả sẽ lưu tại: {result_path}")   # ("danhsach", ".xlsx")

                # --- Khởi tạo trình duyệt và đăng nhập ---
                logger.log("🚀 Khởi động Chrome…")
                options = Options()
                options.add_argument("--start-maximized")
                options.add_experimental_option("detach", True)
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=options)
                wait = WebDriverWait(driver, 20)

                driver.get(base_url)
                logger.log(f"🌐 Mở trang: {base_url}")

                wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(username)
                driver.find_element(By.NAME, "password").send_keys(password)
                driver.find_element(By.NAME, "password").send_keys(Keys.ENTER)
                logger.log("🔐 Đang đăng nhập…")
                messagebox.showinfo("Xác minh",
                                    "Nếu có xác minh thủ công (captcha/SSO), hãy hoàn tất trên trình duyệt rồi bấm OK để tiếp tục.")

                logger.log(f"✅ Đăng nhập thành công. Bắt đầu chọn xã có mã: {ma_xa}")
                option_xpath = f"//select[@id='ddlPhuongXaKeKhai']/option[@value='{ma_xa}']"
                option_element = wait.until(EC.element_to_be_clickable((By.XPATH, option_xpath)))
                option_element.click()
                logger.log(f"✅ Đã chọn xã có mã: {ma_xa}.")

                # --- Mở modal tra cứu MỘT LẦN ---
                logger.log("🔎 Mở cửa sổ tra cứu…")
                tra_cuu_button = wait.until(EC.element_to_be_clickable((By.ID, "btnChonDonDangKy")))
                try:
                    tra_cuu_button.click()
                except ElementClickInterceptedException:
                    logger.log("⚠️ Click bị chặn, thử lại bằng JavaScript...")
                    driver.execute_script("arguments[0].click();", tra_cuu_button)
                wait_tracuu_module_ready(driver, timeout=60)

                # --- Lặp qua từng thửa đất ---
                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                for i, (row_num, so_to, so_thua) in enumerate(plots_to_process):
                    logger.log(f"--- Xử lý thửa {i+1}/{len(plots_to_process)}: Tờ {so_to}, Thửa {so_thua} (Dòng {row_num}) ---")

                    processed, note = search_and_process_plot(driver, wait, logger, so_to, so_thua)
                    logger.log(f"📌 Ghi chú kết quả: {note}")

                    # Ghi vào file Excel KẾT QUẢ
                    result_ws.append([i+1, row_num, so_to, so_thua, note])

                    # Tô màu dòng sau khi xử lý trong file gốc
                    logger.log(f"🎨 Tô màu dòng {row_num} trong file Excel.")
                    for cell in sheet[row_num]:
                        cell.fill = yellow_fill

                    # 💾 Lưu file gốc mỗi 50 dòng
                    if (i + 1) % 50 == 0:
                        try:
                            workbook.save(file_path)
                            logger.log(f"💾 Đã lưu file gốc sau khi xử lý {i+1} dòng.")
                        except Exception as save_err:
                            logger.log(f"⚠️ Lỗi khi lưu file Excel gốc: {save_err}")

                        # Lưu file kết quả mỗi 50 bản ghi
                        try:
                            result_wb.save(result_path)
                            logger.log(f"💾 Đã lưu file kết quả sau {i+1} thửa: {result_path}")
                        except Exception as save_err:
                            logger.log(f"⚠️ Lỗi khi lưu file Excel kết quả: {save_err}")

                    # Nếu đã xử lý (xóa/bỏ) và modal tra cứu đã đóng, cần mở lại
                    if processed:
                        logger.log("🔄 Mở lại cửa sổ tra cứu cho thửa tiếp theo...")
                        tra_cuu_button = wait.until(EC.element_to_be_clickable((By.ID, "btnChonDonDangKy")))
                        try:
                            tra_cuu_button.click()
                        except ElementClickInterceptedException:
                            logger.log("⚠️ Click bị chặn khi mở lại, thử lại bằng JavaScript...")
                            driver.execute_script("arguments[0].click();", tra_cuu_button)
                        wait_tracuu_module_ready(driver, timeout=60)

                # Sau khi xong hết vòng lặp, lưu lần cuối
                try:
                    workbook.save(file_path)
                    logger.log("✅ Đã lưu file Excel gốc lần cuối sau khi hoàn tất toàn bộ.")
                except Exception as save_err:
                    logger.log(f"⚠️ Lỗi khi lưu file Excel gốc lần cuối: {save_err}")

                try:
                    result_wb.save(result_path)
                    logger.log(f"✅ Đã lưu file Excel KẾT QUẢ lần cuối: {result_path}")
                except Exception as save_err:
                    logger.log(f"⚠️ Lỗi khi lưu file Excel KẾT QUẢ lần cuối: {save_err}")
                
                logger.log("✅✅✅ HOÀN TẤT TOÀN BỘ QUÁ TRÌNH! ✅✅✅")

            except Exception as e:
                logger.log(f"❌ Lỗi nghiêm trọng trong quá trình chạy: {e}")
                logger.log(traceback.format_exc())
            finally:
                if driver:
                    logger.log("Trình duyệt vẫn mở. Đóng chương trình để thoát.")
                btn_run.config(state="normal")

        threading.Thread(target=runner, daemon=True).start()

    btn_browse.configure(command=select_excel_file)
    btn_run.configure(command=on_run)
    root.mainloop()

if __name__ == "__main__":
    main()
