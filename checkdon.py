import os
import threading
import traceback
import unicodedata

import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchWindowException

from xoadon_fix import (
    UILogger,
    close_blocking_jconfirm_vbdlis,
    get_login_fields,
    row_is_highlighted,
    wait_all_jconfirm_closed,
    wait_and_count_tblTraCuu,
    wait_query_done,
    wait_tracuu_module_ready,
    wait_tracuu_section_ready,
)


def normalize_header(value):
    return str(value).strip().lower() if value is not None else ""


def normalize_plot_type(value):
    text = str(value or "").strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )
    text = text.replace("đ", "d")
    if text in {"cu", "c", "old", "so cu", "thua cu", "to cu"}:
        return "cu"
    return "moi"


def set_input_value(driver, element, value):
    element.clear()
    element.send_keys(value)
    driver.execute_script(
        """
        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
        """,
        element,
    )


def search_plot_has_don(driver, wait, logger, so_to, so_thua, loai_thua):
    """Tra cuu so to/so thua va tra ve so ban ghi don dang ky tim thay."""
    try:
        wait_all_jconfirm_closed(driver, timeout=10)

        field_names = ("soThuTuThua", "soHieuToBanDo") if loai_thua == "moi" else ("soThuTuThuaCu", "soHieuToBanDoCu")
        for name in ("soThuTuThua", "soHieuToBanDo", "soThuTuThuaCu", "soHieuToBanDoCu"):
            field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"#dvTraCuuTinhHinhDangKyChiTiet input[name='{name}']")))
            set_input_value(driver, field, "")

        so_thua_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"#dvTraCuuTinhHinhDangKyChiTiet input[name='{field_names[0]}']")))
        so_to_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"#dvTraCuuTinhHinhDangKyChiTiet input[name='{field_names[1]}']")))
        set_input_value(driver, so_thua_input, so_thua)
        set_input_value(driver, so_to_input, so_to)

        so_to_input.send_keys(Keys.ENTER)
        wait_query_done(driver)
        wait_tracuu_section_ready(driver)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "tblTraCuuTinhHinhDangKy_info"))
        )

        total_records = wait_and_count_tblTraCuu(driver)
        if total_records == 0:
            logger.log(f"To {so_to}, thua {so_thua} ({loai_thua}): Khong co don.")
            return True, 0, "Khong co don"

        logger.log(f"To {so_to}, thua {so_thua} ({loai_thua}): Co don ({total_records} ban ghi).")
        return True, total_records, f"Co don ({total_records} ban ghi)"

    except NoSuchWindowException:
        logger.log("Cua so trinh duyet da bi dong.")
        return False, None, "Loi: cua so trinh duyet da dong"
    except Exception as exc:
        logger.log(f"Loi khi kiem tra to {so_to}, thua {so_thua}: {exc}")
        logger.log(traceback.format_exc())
        return False, None, f"Loi khi kiem tra: {exc}"


def main():
    root = tk.Tk()
    root.title("Kiem tra thua dat co don hay khong - MPLIS")
    root.geometry("800x650")

    excel_file_path = tk.StringVar()

    main_frm = ttk.Frame(root, padding=12)
    main_frm.pack(fill="both", expand=True)
    main_frm.columnconfigure(1, weight=1)

    ttk.Label(main_frm, text="Username:").grid(row=0, column=0, sticky="e", padx=4, pady=4)
    ent_user = ttk.Entry(main_frm, width=40)
    ent_user.grid(row=0, column=1, sticky="ew", padx=4, pady=4)

    ttk.Label(main_frm, text="Password:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
    ent_pass = ttk.Entry(main_frm, width=40, show="*")
    ent_pass.grid(row=1, column=1, sticky="ew", padx=4, pady=4)

    ttk.Label(main_frm, text="Tinh/Thanh pho:").grid(row=2, column=0, sticky="e", padx=4, pady=4)
    province_cb = ttk.Combobox(main_frm, state="readonly", width=37)
    province_cb["values"] = ["Dak Lak", "Phu Yen"]
    province_cb.grid(row=2, column=1, sticky="w", padx=4, pady=4)
    province_cb.set("Dak Lak")

    ttk.Label(main_frm, text="Ma xa:").grid(row=3, column=0, sticky="e", padx=4, pady=4)
    ent_ma_xa = ttk.Entry(main_frm, width=40)
    ent_ma_xa.grid(row=3, column=1, sticky="w", padx=4, pady=4)

    excel_frm = ttk.LabelFrame(main_frm, text="Cau hinh Excel", padding=10)
    excel_frm.grid(row=4, column=0, columnspan=2, sticky="ew", padx=4, pady=10)
    excel_frm.columnconfigure(1, weight=1)

    btn_browse = ttk.Button(excel_frm, text="Chon file Excel")
    btn_browse.grid(row=0, column=0, padx=4, pady=4)
    ttk.Label(excel_frm, textvariable=excel_file_path, relief="sunken", padding=2).grid(
        row=0, column=1, columnspan=3, sticky="ew", padx=4, pady=4
    )

    ttk.Label(excel_frm, text="Ten cot So to:").grid(row=1, column=0, sticky="e", padx=4, pady=4)
    ent_col_so_to = ttk.Entry(excel_frm, width=20)
    ent_col_so_to.grid(row=1, column=1, sticky="w", padx=4, pady=4)
    ent_col_so_to.insert(0, "soto")

    ttk.Label(excel_frm, text="Ten cot So thua:").grid(row=1, column=2, sticky="e", padx=4, pady=4)
    ent_col_so_thua = ttk.Entry(excel_frm, width=20)
    ent_col_so_thua.grid(row=1, column=3, sticky="w", padx=4, pady=4)
    ent_col_so_thua.insert(0, "sothua")

    ttk.Label(excel_frm, text="Ten cot Moi/Cu:").grid(row=2, column=0, sticky="e", padx=4, pady=4)
    ent_col_loai_thua = ttk.Entry(excel_frm, width=20)
    ent_col_loai_thua.grid(row=2, column=1, sticky="w", padx=4, pady=4)
    ent_col_loai_thua.insert(0, "loai")

    btn_run = ttk.Button(main_frm, text="Chay kiem tra")
    btn_run.grid(row=5, column=1, sticky="w", padx=4, pady=8)

    log_frm = ttk.Frame(main_frm)
    log_frm.grid(row=6, column=0, columnspan=2, sticky="nsew")
    log_frm.columnconfigure(0, weight=1)
    log_frm.rowconfigure(0, weight=1)
    main_frm.rowconfigure(6, weight=1)

    txt = tk.Text(log_frm, state="disabled", bg="#0f1115", fg="#e5e7eb", height=15)
    txt.grid(row=0, column=0, sticky="nsew")
    scrollbar = ttk.Scrollbar(log_frm, orient="vertical", command=txt.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")
    txt["yscrollcommand"] = scrollbar.set

    log = UILogger(txt)

    def select_excel_file():
        filepath = filedialog.askopenfilename(
            title="Chon file Excel",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
        )
        if filepath:
            excel_file_path.set(filepath)

    def on_run():
        username = ent_user.get().strip()
        password = ent_pass.get()
        province = province_cb.get()
        ma_xa = ent_ma_xa.get().strip()
        file_path = excel_file_path.get()
        col_so_to = ent_col_so_to.get().strip().lower()
        col_so_thua = ent_col_so_thua.get().strip().lower()
        col_loai_thua = ent_col_loai_thua.get().strip().lower()

        if not all([username, password, province, ma_xa]):
            messagebox.showerror("Thieu thong tin", "Vui long nhap du Username, Password, Tinh va Ma xa.")
            return
        if not file_path or not col_so_to or not col_so_thua or not col_loai_thua:
            messagebox.showerror("Thieu thong tin Excel", "Vui long chon file Excel va nhap ten cac cot.")
            return

        base_url = "https://phy.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2" if province == "Phu Yen" else "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"
        btn_run.config(state="disabled")

        def runner():
            driver = None
            try:
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                headers = [normalize_header(cell.value) for cell in sheet[1]]

                required_headers = [col_so_to, col_so_thua, col_loai_thua]
                missing_headers = [header for header in required_headers if header not in headers]
                if missing_headers:
                    log.log(f"Khong tim thay cot {missing_headers}. Cac cot hien co: {headers}")
                    return

                idx_so_to = headers.index(col_so_to)
                idx_so_thua = headers.index(col_so_thua)
                idx_loai_thua = headers.index(col_loai_thua)

                plots = []
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    if row_is_highlighted(row):
                        continue
                    so_to_val = row[idx_so_to].value
                    so_thua_val = row[idx_so_thua].value
                    loai_thua_val = row[idx_loai_thua].value
                    if so_to_val and so_thua_val:
                        plots.append((row_idx, str(so_to_val).strip(), str(so_thua_val).strip(), normalize_plot_type(loai_thua_val)))

                if not plots:
                    log.log("Khong co dong nao de kiem tra.")
                    return

                file_root, _ = os.path.splitext(os.path.basename(file_path))
                result_path = os.path.join(os.path.dirname(file_path), f"{ma_xa}_{file_root}_kiem_tra_don.xlsx")

                result_wb = openpyxl.Workbook()
                result_ws = result_wb.active
                result_ws.title = "Ket_qua"
                result_ws.append(["STT", "Dong Excel", "So to", "So thua", "Loai", "So ban ghi", "Ket qua"])

                log.log(f"File ket qua moi: {result_path}")
                log.log("Khoi dong Chrome...")
                options = webdriver.ChromeOptions()
                options.add_argument("--start-maximized")
                options.add_experimental_option("detach", True)
                driver = webdriver.Chrome(options=options)
                wait = WebDriverWait(driver, 20)

                driver.get(base_url)
                log.log(f"Mo trang: {base_url}")

                username_box, password_box = get_login_fields(wait)
                username_box.send_keys(username)
                password_box.send_keys(password)
                password_box.send_keys(Keys.ENTER)
                messagebox.showinfo(
                    "Xac minh",
                    "Neu co captcha/SSO, hay hoan tat tren trinh duyet roi bam OK de tiep tuc.",
                )

                option_xpath = f"//select[@id='ddlPhuongXaKeKhai']/option[@value='{ma_xa}']"
                wait.until(EC.element_to_be_clickable((By.XPATH, option_xpath))).click()
                log.log(f"Da chon ma xa: {ma_xa}")

                tra_cuu_button = wait.until(EC.element_to_be_clickable((By.ID, "btnChonDonDangKy")))
                try:
                    tra_cuu_button.click()
                except ElementClickInterceptedException:
                    if close_blocking_jconfirm_vbdlis(driver, timeout=5):
                        tra_cuu_button = wait.until(EC.element_to_be_clickable((By.ID, "btnChonDonDangKy")))
                    driver.execute_script("arguments[0].click();", tra_cuu_button)
                wait_tracuu_module_ready(driver, timeout=60)

                for i, (row_num, so_to, so_thua, loai_thua) in enumerate(plots, start=1):
                    log.log(f"Kiem tra {i}/{len(plots)}: to {so_to}, thua {so_thua}, loai {loai_thua}, dong Excel {row_num}")
                    ok, total_records, note = search_plot_has_don(driver, wait, log, so_to, so_thua, loai_thua)
                    result_ws.append([i, row_num, so_to, so_thua, loai_thua, total_records, note])

                    if i % 50 == 0:
                        result_wb.save(result_path)
                        log.log(f"Da luu tam sau {i} dong.")

                    if not ok:
                        log.log("Dong nay bi loi, tiep tuc dong ke tiep.")

                result_wb.save(result_path)
                log.log(f"Hoan tat. Da luu file ket qua: {result_path}")
                messagebox.showinfo("Hoan tat", f"Da luu file ket qua:\n{result_path}")

            except Exception as exc:
                log.log(f"Loi nghiem trong: {exc}")
                log.log(traceback.format_exc())
                messagebox.showerror("Loi", str(exc))
            finally:
                if driver:
                    log.log("Trinh duyet van mo. Dong trinh duyet neu muon thoat han.")
                btn_run.config(state="normal")

        threading.Thread(target=runner, daemon=True).start()

    btn_browse.configure(command=select_excel_file)
    btn_run.configure(command=on_run)
    root.mainloop()


if __name__ == "__main__":
    main()
